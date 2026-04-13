from __future__ import annotations

import io
import uuid
from datetime import datetime, time, timedelta, timezone

import pytest
import pytest_asyncio
from httpx import AsyncClient

from tests.conftest import TestSessionLocal
from app.models.store import Store
from app.models.user import User, UserStoreRole, RoleEnum
from app.models.timesheet import TimeEntry, TimeEntryStatus


@pytest_asyncio.fixture
async def seed_store_and_user():
    """Create a store and user with manager role for timesheet tests."""
    async with TestSessionLocal() as session:
        store = Store(
            name="Victoria Enso Jewel",
            location="Jewel Changi Airport",
            address="#02-234 Jewel Changi Airport",
            business_hours_start=time(10, 0),
            business_hours_end=time(22, 0),
        )
        session.add(store)
        await session.flush()
        await session.refresh(store)

        user = User(
            firebase_uid="test-firebase-uid",
            email="test@example.com",
            full_name="Test User",
        )
        session.add(user)
        await session.flush()
        await session.refresh(user)

        # Give user manager role so they can access store-scoped endpoints
        role = UserStoreRole(
            user_id=user.id,
            store_id=store.id,
            role=RoleEnum.manager,
        )
        session.add(role)
        await session.commit()
        await session.refresh(store)
        await session.refresh(user)
        return store, user


class TestClockIn:
    @pytest.mark.asyncio
    async def test_clock_in(self, client: AsyncClient, seed_store_and_user):
        store, user = seed_store_and_user
        resp = await client.post(
            "/api/timesheets/clock-in",
            json={"store_id": str(store.id), "notes": "Starting shift"},
        )
        assert resp.status_code == 201
        data = resp.json()["data"]
        assert data["user_id"] == str(user.id)
        assert data["store_id"] == str(store.id)
        assert data["clock_out"] is None
        assert data["status"] == "pending"
        assert data["notes"] == "Starting shift"

        # Verify status endpoint shows clocked in
        status_resp = await client.get("/api/timesheets/status")
        assert status_resp.status_code == 200
        status_data = status_resp.json()["data"]
        assert status_data is not None
        assert status_data["id"] == data["id"]

    @pytest.mark.asyncio
    async def test_double_clock_in(self, client: AsyncClient, seed_store_and_user):
        store, _ = seed_store_and_user
        # First clock in
        resp1 = await client.post(
            "/api/timesheets/clock-in",
            json={"store_id": str(store.id)},
        )
        assert resp1.status_code == 201

        # Second clock in should fail
        resp2 = await client.post(
            "/api/timesheets/clock-in",
            json={"store_id": str(store.id)},
        )
        assert resp2.status_code == 400
        assert "Already clocked in" in resp2.json()["detail"]


class TestClockOut:
    @pytest.mark.asyncio
    async def test_clock_out(self, client: AsyncClient, seed_store_and_user):
        store, _ = seed_store_and_user
        # Clock in first
        await client.post(
            "/api/timesheets/clock-in",
            json={"store_id": str(store.id)},
        )

        # Clock out
        resp = await client.post(
            "/api/timesheets/clock-out",
            json={"break_minutes": 30, "notes": "End of shift"},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["clock_out"] is not None
        assert data["break_minutes"] == 30
        assert data["hours_worked"] is not None
        # hours_worked should be >= 0
        assert data["hours_worked"] >= 0

    @pytest.mark.asyncio
    async def test_clock_out_without_clock_in(self, client: AsyncClient, seed_store_and_user):
        # Don't clock in, just try to clock out
        resp = await client.post(
            "/api/timesheets/clock-out",
            json={},
        )
        assert resp.status_code == 400
        assert "Not currently clocked in" in resp.json()["detail"]


class TestListTimesheets:
    @pytest.mark.asyncio
    async def test_list_timesheets(self, client: AsyncClient, seed_store_and_user):
        store, user = seed_store_and_user

        # Create some time entries directly in DB
        now = datetime.now(timezone.utc)
        async with TestSessionLocal() as session:
            entry1 = TimeEntry(
                user_id=user.id,
                store_id=store.id,
                clock_in=now - timedelta(hours=10),
                clock_out=now - timedelta(hours=2),
                break_minutes=30,
                status=TimeEntryStatus.pending,
            )
            entry2 = TimeEntry(
                user_id=user.id,
                store_id=store.id,
                clock_in=now - timedelta(days=2, hours=10),
                clock_out=now - timedelta(days=2, hours=2),
                break_minutes=0,
                status=TimeEntryStatus.pending,
            )
            session.add_all([entry1, entry2])
            await session.commit()

        # List all
        resp = await client.get(f"/api/stores/{store.id}/timesheets")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 2

        # Filter by date_from (only recent entry)
        date_from = (now - timedelta(hours=12)).strftime("%Y-%m-%dT%H:%M:%S")
        resp2 = await client.get(
            f"/api/stores/{store.id}/timesheets?date_from={date_from}"
        )
        assert resp2.status_code == 200
        assert resp2.json()["total"] == 1


class TestApproveTimesheet:
    @pytest.mark.asyncio
    async def test_approve_timesheet(self, client: AsyncClient, seed_store_and_user):
        store, user = seed_store_and_user

        now = datetime.now(timezone.utc)
        async with TestSessionLocal() as session:
            entry = TimeEntry(
                user_id=user.id,
                store_id=store.id,
                clock_in=now - timedelta(hours=8),
                clock_out=now,
                break_minutes=30,
                status=TimeEntryStatus.pending,
            )
            session.add(entry)
            await session.commit()
            await session.refresh(entry)
            entry_id = entry.id

        resp = await client.patch(
            f"/api/stores/{store.id}/timesheets/{entry_id}",
            json={"status": "approved"},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["status"] == "approved"
        assert data["approved_by"] == str(user.id)



class TestImportTimesheets:
    """Tests for POST /api/stores/{store_id}/timesheets/import."""

    @pytest_asyncio.fixture
    async def seed_staff(self, seed_store_and_user):
        """Create a second staff user to import timesheets for."""
        store, manager = seed_store_and_user
        async with TestSessionLocal() as session:
            staff = User(
                firebase_uid="staff-firebase-uid",
                email="staff@example.com",
                full_name="Jane Doe",
            )
            session.add(staff)
            await session.flush()
            role = UserStoreRole(
                user_id=staff.id,
                store_id=store.id,
                role=RoleEnum.staff,
            )
            session.add(role)
            await session.commit()
            await session.refresh(staff)
            return store, manager, staff

    @pytest.mark.asyncio
    async def test_import_csv_success(self, client: AsyncClient, seed_staff):
        store, manager, staff = seed_staff
        csv_content = (
            "email,date,clock_in_time,clock_out_time,break_minutes,notes\n"
            f"{staff.email},2026-01-05,09:00,17:30,30,Morning shift\n"
            f"{manager.email},2026-01-05,10:00,18:00,0,\n"
        )
        resp = await client.post(
            f"/api/stores/{store.id}/timesheets/import",
            files={"file": ("timesheets.csv", csv_content.encode(), "text/csv")},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported_count"] == 2
        assert data["skipped_count"] == 0
        assert data["errors"] == []

    @pytest.mark.asyncio
    async def test_import_csv_by_name(self, client: AsyncClient, seed_staff):
        store, _, staff = seed_staff
        csv_content = (
            "staff_name,date,clock_in_time,clock_out_time,break_minutes,notes\n"
            f"{staff.full_name},2026-02-10,08:00,16:00,30,\n"
        )
        resp = await client.post(
            f"/api/stores/{store.id}/timesheets/import",
            files={"file": ("import.csv", csv_content.encode(), "text/csv")},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported_count"] == 1
        assert data["skipped_count"] == 0

    @pytest.mark.asyncio
    async def test_import_csv_unknown_staff(self, client: AsyncClient, seed_store_and_user):
        store, _ = seed_store_and_user
        csv_content = (
            "email,date,clock_in_time,clock_out_time,break_minutes,notes\n"
            "nobody@example.com,2026-01-05,09:00,17:00,0,\n"
        )
        resp = await client.post(
            f"/api/stores/{store.id}/timesheets/import",
            files={"file": ("import.csv", csv_content.encode(), "text/csv")},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported_count"] == 0
        assert data["skipped_count"] == 1
        assert len(data["errors"]) == 1
        assert "Staff not found" in data["errors"][0]["message"]

    @pytest.mark.asyncio
    async def test_import_csv_invalid_time(self, client: AsyncClient, seed_staff):
        store, _, staff = seed_staff
        csv_content = (
            "email,date,clock_in_time,clock_out_time,break_minutes,notes\n"
            f"{staff.email},2026-01-05,BADTIME,17:00,0,\n"
        )
        resp = await client.post(
            f"/api/stores/{store.id}/timesheets/import",
            files={"file": ("import.csv", csv_content.encode(), "text/csv")},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported_count"] == 0
        assert data["skipped_count"] == 1
        assert "Invalid clock-in time" in data["errors"][0]["message"]

    @pytest.mark.asyncio
    async def test_import_csv_duplicate_entry(self, client: AsyncClient, seed_staff):
        store, _, staff = seed_staff
        csv_content = (
            "email,date,clock_in_time,clock_out_time,break_minutes,notes\n"
            f"{staff.email},2026-03-01,09:00,17:00,30,\n"
        )
        # First import
        resp1 = await client.post(
            f"/api/stores/{store.id}/timesheets/import",
            files={"file": ("import.csv", csv_content.encode(), "text/csv")},
        )
        assert resp1.json()["data"]["imported_count"] == 1

        # Second import with same user+date should skip as duplicate
        resp2 = await client.post(
            f"/api/stores/{store.id}/timesheets/import",
            files={"file": ("import.csv", csv_content.encode(), "text/csv")},
        )
        data = resp2.json()["data"]
        assert data["imported_count"] == 0
        assert data["skipped_count"] == 1
        assert "Duplicate" in data["errors"][0]["message"]

    @pytest.mark.asyncio
    async def test_import_csv_clock_out_before_in(self, client: AsyncClient, seed_staff):
        store, _, staff = seed_staff
        csv_content = (
            "email,date,clock_in_time,clock_out_time,break_minutes,notes\n"
            f"{staff.email},2026-01-05,17:00,09:00,0,\n"
        )
        resp = await client.post(
            f"/api/stores/{store.id}/timesheets/import",
            files={"file": ("import.csv", csv_content.encode(), "text/csv")},
        )
        data = resp.json()["data"]
        assert data["imported_count"] == 0
        assert data["skipped_count"] == 1
        assert "Clock-out must be after clock-in" in data["errors"][0]["message"]

    @pytest.mark.asyncio
    async def test_import_creates_approved_entries(self, client: AsyncClient, seed_staff):
        """Imported entries should have status=approved."""
        store, _, staff = seed_staff
        csv_content = (
            "email,date,clock_in_time,clock_out_time,break_minutes,notes\n"
            f"{staff.email},2026-04-01,09:00,17:00,30,Historical\n"
        )
        await client.post(
            f"/api/stores/{store.id}/timesheets/import",
            files={"file": ("import.csv", csv_content.encode(), "text/csv")},
        )
        # Check via list endpoint
        resp = await client.get(
            f"/api/stores/{store.id}/timesheets?status=approved"
        )
        assert resp.status_code == 200
        entries = resp.json()["data"]
        assert len(entries) >= 1
        assert entries[0]["status"] == "approved"

    @pytest.mark.asyncio
    async def test_import_unsupported_format(self, client: AsyncClient, seed_store_and_user):
        store, _ = seed_store_and_user
        resp = await client.post(
            f"/api/stores/{store.id}/timesheets/import",
            files={"file": ("data.txt", b"some text", "text/plain")},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported_count"] == 0
        assert "Unsupported file format" in data["errors"][0]["message"]


# ---------------------------------------------------------------------------
# VE Payroll Import Tests
# ---------------------------------------------------------------------------

def _build_ve_payroll_xlsx(
    sheet_name: str = "Jan 2025",
    staff: list = None,
    daily_data: list = None,
) -> bytes:
    """Build a synthetic VE payroll .xlsx in memory.

    staff: list of dicts with keys: name, col, total_hrs, hour_rate, sales_pct,
           total_sales, gross_hour, commission, gross_payment
    daily_data: list of (date, col, hours, sales) tuples
    """
    from openpyxl import Workbook

    if staff is None:
        staff = [
            {
                "name": "Test User",
                "col": 2,  # B
                "total_hrs": 20,
                "hour_rate": 11,
                "sales_pct": 10,
                "total_sales": 5000,
                "gross_hour": 220,
                "commission": 500,
                "gross_payment": 720,
            },
        ]
    if daily_data is None:
        daily_data = [
            (datetime(2025, 1, 2), 2, 10, 2500),
            (datetime(2025, 1, 3), 2, 10, 2500),
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Row 1: title
    ws["A1"] = f"{sheet_name} Payroll"

    # Row 12: staff "Person: XXX"
    for s in staff:
        ws.cell(row=12, column=s["col"], value=f"Person: {s['name']}")

    # Summary rows
    labels = [
        (14, "Total Hrs", "total_hrs"),
        (17, "Hour Rate", "hour_rate"),
        (18, "Sales (%)", "sales_pct"),
        (16, "Total Sales", "total_sales"),
        (19, "Gross Hour", "gross_hour"),
        (20, "Commission", "commission"),
        (21, "Gross Payment", "gross_payment"),
    ]
    for row, label, key in labels:
        ws.cell(row=row, column=1, value=label)
        for s in staff:
            ws.cell(row=row, column=s["col"], value=s.get(key, 0))

    # Row 28: headers
    ws.cell(row=28, column=1, value="Date")
    for s in staff:
        ws.cell(row=28, column=s["col"], value="Hrs")
        ws.cell(row=28, column=s["col"] + 1, value="Sales ($)")

    # Daily data rows
    for i, (dt, col, hrs, sales) in enumerate(daily_data):
        row = 29 + i
        ws.cell(row=row, column=1, value=dt)
        ws.cell(row=row, column=col, value=hrs)
        ws.cell(row=row, column=col + 1, value=sales)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class TestVEPayrollImport:
    """Tests for POST /api/stores/{store_id}/timesheets/import-ve-payroll."""

    @pytest.mark.asyncio
    async def test_ve_import_creates_entries_and_orders(
        self, client: AsyncClient, seed_store_and_user
    ):
        store, user = seed_store_and_user
        xlsx = _build_ve_payroll_xlsx(
            staff=[{
                "name": user.full_name,
                "col": 2,
                "total_hrs": 20,
                "hour_rate": 11,
                "sales_pct": 10,
                "total_sales": 5000,
                "gross_hour": 220,
                "commission": 500,
                "gross_payment": 720,
            }],
            daily_data=[
                (datetime(2025, 1, 2), 2, 10.5, 2500),
                (datetime(2025, 1, 3), 2, 9.5, 2500),
            ],
        )
        resp = await client.post(
            f"/api/stores/{store.id}/timesheets/import-ve-payroll",
            files={"file": ("payroll.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert len(data["months"]) == 1
        month = data["months"][0]
        assert month["sheet_name"] == "Jan 2025"
        assert len(month["staff"]) == 1
        s = month["staff"][0]
        assert s["name"] == user.full_name
        assert s["time_entries_created"] == 2
        assert s["orders_created"] == 2
        assert s["total_hours"] == 20

    @pytest.mark.asyncio
    async def test_ve_import_creates_user_if_missing(
        self, client: AsyncClient, seed_store_and_user
    ):
        store, _ = seed_store_and_user
        xlsx = _build_ve_payroll_xlsx(
            staff=[{
                "name": "NewPerson",
                "col": 2,
                "total_hrs": 8,
                "hour_rate": 10,
                "sales_pct": 10,
                "total_sales": 1000,
                "gross_hour": 80,
                "commission": 100,
                "gross_payment": 180,
            }],
            daily_data=[
                (datetime(2025, 2, 1), 2, 8, 1000),
            ],
        )
        resp = await client.post(
            f"/api/stores/{store.id}/timesheets/import-ve-payroll",
            files={"file": ("payroll.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        month = data["months"][0]
        assert month["staff"][0]["time_entries_created"] == 1
        # Profile should have been created
        assert "NewPerson" in data["profiles_updated"]

    @pytest.mark.asyncio
    async def test_ve_import_skips_duplicates(
        self, client: AsyncClient, seed_store_and_user
    ):
        store, user = seed_store_and_user
        xlsx = _build_ve_payroll_xlsx(
            staff=[{
                "name": user.full_name,
                "col": 2,
                "total_hrs": 10,
                "hour_rate": 11,
                "sales_pct": 10,
                "total_sales": 2000,
                "gross_hour": 110,
                "commission": 200,
                "gross_payment": 310,
            }],
            daily_data=[
                (datetime(2025, 3, 1), 2, 10, 2000),
            ],
        )
        # First import
        resp1 = await client.post(
            f"/api/stores/{store.id}/timesheets/import-ve-payroll",
            files={"file": ("payroll.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp1.json()["data"]["months"][0]["staff"][0]["time_entries_created"] == 1

        # Second import — should skip duplicates
        resp2 = await client.post(
            f"/api/stores/{store.id}/timesheets/import-ve-payroll",
            files={"file": ("payroll.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        s = resp2.json()["data"]["months"][0]["staff"][0]
        assert s["time_entries_created"] == 0
        assert s["time_entries_skipped"] == 1

    @pytest.mark.asyncio
    async def test_ve_import_multiple_staff(
        self, client: AsyncClient, seed_store_and_user
    ):
        store, user = seed_store_and_user
        xlsx = _build_ve_payroll_xlsx(
            staff=[
                {
                    "name": user.full_name,
                    "col": 2,
                    "total_hrs": 10,
                    "hour_rate": 12,
                    "sales_pct": 10,
                    "total_sales": 3000,
                    "gross_hour": 120,
                    "commission": 300,
                    "gross_payment": 420,
                },
                {
                    "name": "AnotherStaff",
                    "col": 7,
                    "total_hrs": 8,
                    "hour_rate": 10,
                    "sales_pct": 10,
                    "total_sales": 1000,
                    "gross_hour": 80,
                    "commission": 100,
                    "gross_payment": 180,
                },
            ],
            daily_data=[
                (datetime(2025, 4, 1), 2, 10, 3000),
                (datetime(2025, 4, 1), 7, 8, 1000),
            ],
        )
        resp = await client.post(
            f"/api/stores/{store.id}/timesheets/import-ve-payroll",
            files={"file": ("payroll.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        staff_list = data["months"][0]["staff"]
        assert len(staff_list) == 2
        names = {s["name"] for s in staff_list}
        assert user.full_name in names
        assert "AnotherStaff" in names