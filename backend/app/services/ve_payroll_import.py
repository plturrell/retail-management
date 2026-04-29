"""Service for importing Victoria Enso legacy payroll Excel files.

Format: multi-sheet workbook where each sheet = one month.
- Row 12-13: "Person: XXX" cells identify staff columns
- Rows 14-27: monthly summary (total hrs, rates, gross, CPF, etc.)
- Row 28-29: column headers (Date, Hrs, sales channels...)
- Row 29/30+: daily data (datetime date, hours, sales amounts)
"""
from __future__ import annotations

import re
import uuid as _uuid
from datetime import date, datetime, time, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

import io
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from google.cloud.firestore_v1.client import Client as FirestoreClient

import app.firestore as _fs  # access as `_fs.db` so init stays lazy
from app.firestore_helpers import query_collection, create_document, get_document, update_document


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

class StaffColumn:
    """Detected staff person and their column position."""
    def __init__(self, name: str, col_idx: int, row: int):
        self.name = name          # e.g. "Evonne"
        self.col_idx = col_idx    # 1-based column index (B=2, G=7, ...)
        self.row = row            # row where "Person: XXX" was found


class MonthlyStaffData:
    """Parsed monthly data for one staff member."""
    def __init__(self, name: str):
        self.name = name
        self.total_hours: float = 0.0
        self.hour_rate: float = 0.0
        self.sales_pct: float = 0.0
        self.total_sales: float = 0.0
        self.gross_hour: float = 0.0
        self.commission: float = 0.0
        self.gross_payment: float = 0.0
        self.other_payment: float = 0.0
        self.giro_payment: float = 0.0
        # Daily entries: list of (date, hours, total_daily_sales)
        self.daily_entries: List[Tuple[date, float, float]] = []


class MonthImportResult:
    """Result from importing one sheet/month."""
    def __init__(self, sheet_name: str):
        self.sheet_name = sheet_name
        self.staff_results: List[StaffImportResult] = []
        self.errors: List[Dict[str, Any]] = []


class StaffImportResult:
    def __init__(self, name: str):
        self.name = name
        self.time_entries_created: int = 0
        self.time_entries_skipped: int = 0
        self.orders_created: int = 0
        self.total_hours: float = 0.0
        self.total_sales: float = 0.0


class VEImportResult:
    """Aggregate result across all sheets."""
    def __init__(self):
        self.months: List[MonthImportResult] = []
        self.users_created: List[str] = []
        self.profiles_updated: List[str] = []
        self.errors: List[Dict[str, Any]] = []

    def to_dict(self) -> dict:
        return {
            "months": [
                {
                    "sheet_name": m.sheet_name,
                    "staff": [
                        {
                            "name": s.name,
                            "time_entries_created": s.time_entries_created,
                            "time_entries_skipped": s.time_entries_skipped,
                            "orders_created": s.orders_created,
                            "total_hours": s.total_hours,
                            "total_sales": s.total_sales,
                        }
                        for s in m.staff_results
                    ],
                    "errors": m.errors,
                }
                for m in self.months
            ],
            "users_created": self.users_created,
            "profiles_updated": self.profiles_updated,
            "errors": self.errors,
        }


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

_PERSON_RE = re.compile(r"Person:\s*(.+)", re.IGNORECASE)


def _detect_staff_columns(ws: Worksheet) -> List[StaffColumn]:
    """Scan rows 12-13 for 'Person: XXX' cells to locate staff columns."""
    staff: List[StaffColumn] = []
    for row_idx in (12, 13):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and isinstance(val, str):
                m = _PERSON_RE.match(val.strip())
                if m:
                    staff.append(StaffColumn(m.group(1).strip(), col_idx, row_idx))
    return staff


def _find_label_row(ws: Worksheet, label: str, col: int = 1,
                    start: int = 13, end: int = 30) -> Optional[int]:
    """Find the row where column `col` contains `label` (case-insensitive)."""
    for r in range(start, end + 1):
        v = ws.cell(row=r, column=col).value
        if v and isinstance(v, str) and label.lower() in v.strip().lower():
            return r
    return None


def _safe_float(val: Any) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _find_header_row(ws: Worksheet) -> Optional[int]:
    """Find the row containing 'Date' in column A (header row for daily data)."""
    for r in range(25, 35):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, str) and v.strip().lower() == "date":
            return r
    return None


def _parse_month_sheet(ws: Worksheet) -> Tuple[List[MonthlyStaffData], List[Dict]]:
    """Parse a single month sheet and return staff data + errors."""
    errors: List[Dict] = []
    staff_cols = _detect_staff_columns(ws)
    if not staff_cols:
        errors.append({"message": f"No 'Person: XXX' cells found in sheet"})
        return [], errors

    # Find key label rows
    total_hrs_row = _find_label_row(ws, "Total Hrs")
    hour_rate_row = _find_label_row(ws, "Hour Rate")
    sales_pct_row = _find_label_row(ws, "Sales (%)")
    total_sales_row = _find_label_row(ws, "Total Sales", start=14)
    gross_hour_row = _find_label_row(ws, "Gross Hour")
    commission_row = _find_label_row(ws, "Commission")
    gross_pay_row = _find_label_row(ws, "Gross Payment")
    other_pay_row = _find_label_row(ws, "Other Payment")
    giro_row = _find_label_row(ws, "Giro Payment")
    header_row = _find_header_row(ws)

    if header_row is None:
        errors.append({"message": "Could not find 'Date' header row"})
        return [], errors

    results: List[MonthlyStaffData] = []

    for sc in staff_cols:
        data = MonthlyStaffData(sc.name)
        c = sc.col_idx

        # Extract monthly summary values
        if total_hrs_row:
            data.total_hours = _safe_float(ws.cell(row=total_hrs_row, column=c).value)
        if hour_rate_row:
            data.hour_rate = _safe_float(ws.cell(row=hour_rate_row, column=c).value)
        if sales_pct_row:
            data.sales_pct = _safe_float(ws.cell(row=sales_pct_row, column=c).value)
        if total_sales_row:
            data.total_sales = _safe_float(ws.cell(row=total_sales_row, column=c).value)
        if gross_hour_row:
            data.gross_hour = _safe_float(ws.cell(row=gross_hour_row, column=c).value)
        if commission_row:
            data.commission = _safe_float(ws.cell(row=commission_row, column=c).value)
        if gross_pay_row:
            data.gross_payment = _safe_float(ws.cell(row=gross_pay_row, column=c).value)
        if other_pay_row:
            data.other_payment = _safe_float(ws.cell(row=other_pay_row, column=c).value)
        if giro_row:
            data.giro_payment = _safe_float(ws.cell(row=giro_row, column=c).value)

        # Parse daily entries starting after header row
        # The hours column for this staff is at sc.col_idx
        # Sales columns follow (col_idx+1, col_idx+2, ...)
        # Determine how many sales columns this person has by looking at
        # the header row: count non-None cells after the Hrs column until next staff or gap
        sales_col_count = 0
        for check_col in range(c + 1, c + 6):
            hdr = ws.cell(row=header_row, column=check_col).value
            if hdr is None or (isinstance(hdr, str) and hdr.strip().lower() == "hrs"):
                break
            sales_col_count += 1

        for r in range(header_row + 1, ws.max_row + 1):
            date_val = ws.cell(row=r, column=1).value
            if date_val is None:
                continue
            if isinstance(date_val, datetime):
                entry_date = date_val.date()
            elif isinstance(date_val, date):
                entry_date = date_val
            else:
                continue

            hours = _safe_float(ws.cell(row=r, column=c).value)
            # Sum sales from all sales columns for this person
            daily_sales = 0.0
            for sc_offset in range(1, sales_col_count + 1):
                daily_sales += _safe_float(ws.cell(row=r, column=c + sc_offset).value)

            if hours > 0 or daily_sales > 0:
                data.daily_entries.append((entry_date, hours, daily_sales))

        results.append(data)

    return results, errors



# ---------------------------------------------------------------------------
# User resolution
# ---------------------------------------------------------------------------

def _resolve_or_create_user(
    name: str, user_lookup: Dict[str, dict],
) -> dict:
    """Find a user by name (case-insensitive) or create a placeholder in Firestore."""
    key = name.lower()
    if key in user_lookup:
        return user_lookup[key]
    # Create placeholder user in Firestore
    email = f"{name.lower().replace(' ', '.')}@victoriaenso.sg"
    doc_id = str(_uuid.uuid4())
    now = datetime.now(timezone.utc)
    user_data = {
        "firebase_uid": f"ve-import-{_uuid.uuid4().hex[:12]}",
        "email": email,
        "full_name": name,
        "phone": None,
        "created_at": now,
        "updated_at": now,
    }
    result = create_document("users", user_data, doc_id=doc_id)
    user_lookup[key] = result
    user_lookup[email.lower()] = result
    return result


def _entry_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).date()
        except ValueError:
            return None
    return None


# ---------------------------------------------------------------------------
# Main import function
# ---------------------------------------------------------------------------

def import_ve_payroll(
    db: FirestoreClient,
    store_id: _uuid.UUID,
    content: bytes,
) -> VEImportResult:
    """Import a VE-format payroll Excel workbook."""
    result = VEImportResult()

    wb = load_workbook(filename=io.BytesIO(content), data_only=True)

    # Build user lookup from Firestore
    all_users = sorted(
        query_collection("users"),
        key=lambda user: (
            str(user.get("updated_at") or ""),
            str(user.get("created_at") or ""),
            str(user.get("id") or ""),
        ),
    )
    user_lookup: Dict[str, dict] = {}
    for u in all_users:
        if u.get("email"):
            user_lookup[u["email"].lower()] = u
        if u.get("full_name"):
            user_lookup[u["full_name"].lower()] = u

    # Skip non-data sheets
    skip_sheets = {"sheet3", "sheet2", "sheet1"}

    timesheet_col = f"stores/{store_id}/timesheets"
    orders_col = f"stores/{store_id}/orders"
    existing_time_entry_keys = {
        (str(entry.get("user_id", "")), entry_date)
        for entry in query_collection(timesheet_col, order_by="-clock_in")
        if (entry_date := _entry_date(entry.get("clock_in"))) is not None
    }

    batch = _fs.db.batch()
    batch_count = 0

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in skip_sheets:
            continue
        ws = wb[sheet_name]
        month_result = MonthImportResult(sheet_name)

        staff_data_list, parse_errors = _parse_month_sheet(ws)
        month_result.errors.extend(parse_errors)

        for staff_data in staff_data_list:
            user = _resolve_or_create_user(staff_data.name, user_lookup)

            staff_result = StaffImportResult(staff_data.name)
            staff_result.total_hours = staff_data.total_hours
            staff_result.total_sales = staff_data.total_sales

            user_id = user.get("id", "")
            user_name = user.get("full_name", staff_data.name)

            # Update/create EmployeeProfile with hourly rate and commission
            if staff_data.hour_rate > 0:
                profiles = query_collection(
                    "employee_profiles",
                    filters=[("user_id", "==", str(user_id))],
                    limit=1,
                )
                now = datetime.now(timezone.utc)
                if not profiles:
                    profile_id = str(_uuid.uuid4())
                    profile_data = {
                        "user_id": str(user_id),
                        "date_of_birth": "1990-01-01",
                        "nationality": "foreigner",
                        "basic_salary": 0,
                        "hourly_rate": staff_data.hour_rate,
                        "commission_rate": staff_data.sales_pct,
                        "start_date": "2024-01-01",
                        "created_at": now,
                        "updated_at": now,
                    }
                    create_document("employee_profiles", profile_data, doc_id=profile_id)
                    if staff_data.name not in result.profiles_updated:
                        result.profiles_updated.append(staff_data.name)
                else:
                    profile = profiles[0]
                    update_fields = {"hourly_rate": staff_data.hour_rate, "updated_at": now}
                    if staff_data.sales_pct > 0:
                        update_fields["commission_rate"] = staff_data.sales_pct
                    update_document("employee_profiles", profile["id"], update_fields)
                    if staff_data.name not in result.profiles_updated:
                        result.profiles_updated.append(staff_data.name)

            # Create daily TimeEntry records
            for entry_date, hours, daily_sales in staff_data.daily_entries:
                if hours > 0:
                    duplicate_key = (str(user_id), entry_date)
                    if duplicate_key in existing_time_entry_keys:
                        staff_result.time_entries_skipped += 1
                    else:
                        clock_in = datetime.combine(entry_date, time(9, 0))
                        clock_out_dt = clock_in + timedelta(hours=hours)
                        doc_id = str(_uuid.uuid4())
                        now = datetime.now(timezone.utc)
                        ref = _fs.db.collection(timesheet_col).document(doc_id)
                        batch.set(ref, {
                            "id": doc_id,
                            "user_id": str(user_id),
                            "store_id": str(store_id),
                            "clock_in": clock_in,
                            "clock_out": clock_out_dt,
                            "break_minutes": 0,
                            "notes": f"VE payroll import: {sheet_name}",
                            "status": "approved",
                            "approved_by": None,
                            "user_name": user_name,
                            "created_at": now,
                            "updated_at": now,
                        })
                        batch_count += 1
                        staff_result.time_entries_created += 1
                        existing_time_entry_keys.add(duplicate_key)

                # Create Order record for daily sales
                if daily_sales > 0:
                    order_num = (
                        f"VE-{entry_date.strftime('%Y%m%d')}"
                        f"-{user_name.replace(' ', '')[:6].upper()}"
                        f"-{_uuid.uuid4().hex[:4]}"
                    )
                    order_doc_id = str(_uuid.uuid4())
                    now = datetime.now(timezone.utc)
                    ref = _fs.db.collection(orders_col).document(order_doc_id)
                    batch.set(ref, {
                        "id": order_doc_id,
                        "order_number": order_num,
                        "store_id": str(store_id),
                        "staff_id": str(user_id),
                        "salesperson_id": str(user_id),
                        "order_date": datetime.combine(entry_date, time(12, 0)),
                        "subtotal": daily_sales,
                        "grand_total": daily_sales,
                        "payment_method": "mixed",
                        "status": "completed",
                        "source": "manual",
                        "created_at": now,
                        "updated_at": now,
                    })
                    batch_count += 1
                    staff_result.orders_created += 1

                # Commit batch if approaching limit
                if batch_count >= 499:
                    batch.commit()
                    batch = _fs.db.batch()
                    batch_count = 0

            month_result.staff_results.append(staff_result)

        result.months.append(month_result)

    if batch_count > 0:
        batch.commit()
    wb.close()
    return result
