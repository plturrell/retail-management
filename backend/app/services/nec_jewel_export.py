"""Helpers for building the Jewel NEC POS master data workbook."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BRAND_NAME = "VICTORIA ENSO"
DEFAULT_INV_STORE_CODE = "JEWEL-01"
DEFAULT_TENANT_CODE = "VE_JEWEL"
TAX_CODE = "G"
GST_RATE = 0.09
FAR_FUTURE = "20991231"

# Backward-compatibility alias. New code should pass tenant_code explicitly
# (typically read from the target store's `tenant_code` field in Firestore).
TENANT_CODE = DEFAULT_TENANT_CODE

PROMO_TIERS: list[tuple[str, str, float]] = [
    ("VE_GENERAL_10", "10% General", 10),
    ("VE_SPECIAL_15", "15% Special", 15),
    ("VE_DIRECTOR_20", "20% Director Approved", 20),
    ("VE_CLEARANCE_25", "25% Clearance Special", 25),
    ("VE_STAFFCARD_20", "20% Jewel Staff Card", 20),
    ("VE_VIP_LOYALTY_20", "20% VIP Loyalty Customer", 20),
]

CAG_CATG_MAP: dict[str, str] = {
    "Bracelet": "BANGLE & BRACELETS",
    "Necklace": "NECKLACE",
    "Ring": "RINGS",
    "Earring": "EARRINGS",
    "Charm": "CHARMS",
    "Pendant": "CHARMS",
    "Bead Strand": "COSTUME JEWELLERY",
    "Accessory": "JEWELLERY ACCESSORY",
    "Loose Gemstone": "PRECIOUS STONE/GOLD",
    "Raw Specimen": "PRECIOUS STONE/GOLD",
    "Crystal Cluster": "PRECIOUS STONE/GOLD",
    "Crystal Point": "PRECIOUS STONE/GOLD",
    "Tumbled Stone": "PRECIOUS STONE/GOLD",
    "Gemstone Bead": "PRECIOUS STONE/GOLD",
    "Healing Crystal": "PRECIOUS STONE/GOLD",
    "Figurine": "DECORATIVE ITEM",
    "Sculpture": "DECORATIVE ITEM",
    "Bookend": "DECORATIVE ITEM",
    "Bowl": "DECORATIVE ITEM",
    "Vase": "DECORATIVE ITEM",
    "Box": "DECORATIVE ITEM",
    "Tray": "DECORATIVE ITEM",
    "Decorative Object": "DECORATIVE ITEM",
    "Wall Art": "DECORATIVE ITEM",
    "Gift Set": "GENERAL SOUVENIRS",
    "Repair Service": "REPAIR SERVICE",
}

def make_tenant_catg_tree(tenant_code: str) -> list[tuple[str, str, str, str]]:
    """Build the parent/child category tree for a given tenant.

    Only the three root rows reference the tenant code; subcategories are
    fixed and tenant-agnostic. Pass the tenant_code from the store's
    Firestore doc (`stores/{id}.tenant_code`).
    """
    return [
        (tenant_code, "VE_JEWELLERY", "Jewellery", ""),
        (tenant_code, "VE_HOMEDECOR", "Home Decor", ""),
        (tenant_code, "VE_MINERALS", "Specimen Minerals", ""),
        ("VE_JEWELLERY", "VE_JW_BRACELET", "Bracelets", "BANGLE & BRACELETS"),
        ("VE_JEWELLERY", "VE_JW_NECKLACE", "Necklaces", "NECKLACE"),
        ("VE_JEWELLERY", "VE_JW_RING", "Rings", "RINGS"),
        ("VE_JEWELLERY", "VE_JW_EARRING", "Earrings", "EARRINGS"),
        ("VE_JEWELLERY", "VE_JW_CHARM", "Charms & Pendants", "CHARMS"),
        ("VE_JEWELLERY", "VE_JW_COSTUME", "Costume Jewellery", "COSTUME JEWELLERY"),
        ("VE_JEWELLERY", "VE_JW_ACC", "Jewellery Accessory", "JEWELLERY ACCESSORY"),
        ("VE_JEWELLERY", "VE_JW_REPAIR", "Repair Service", "REPAIR SERVICE"),
        ("VE_HOMEDECOR", "VE_HD_DECOR", "Decorative Items", "DECORATIVE ITEM"),
        ("VE_HOMEDECOR", "VE_HD_GENERAL", "Gifts & General", "GENERAL SOUVENIRS"),
        ("VE_MINERALS", "VE_SM_STONE", "Precious Stones & Crystals", "PRECIOUS STONE/GOLD"),
    ]


# Backward-compatible default tree using DEFAULT_TENANT_CODE.
TENANT_CATG_TREE = make_tenant_catg_tree(DEFAULT_TENANT_CODE)

_CAG_TO_TENANT: dict[str, str] = {
    cag: child for _parent, child, _desc, cag in TENANT_CATG_TREE if cag
}

SKU_V2_HEADERS = [
    "FILENAME", "MODE", "SKU_CODE", "SKU_DESC", "COSTPRICE", "AVGCOST",
    "SKU_PRICE", "SKU_PRICE1", "SKU-EDATE1", "SKU-ETIME1",
    "SKU_CATG_TENANT", "SKU_CATG_CAG", "SKU_ROL", "SKU_ROQ", "SKU_QOH",
    "TAX_CODE", "PRO-FRDATE", "PRO_FRTIME", "PRO_TODATE", "PRO_TOTIME",
    "PRO_PRICE", "OPEN_ITEM", "KIT", "REMARKS", "SKU_DISC",
    "SKU_PRICE_LALO", "SKU_PRICE_HALO", "MIN_AGE",
    "BRAND", "GENDER", "AGE GROUP", "CHANGI COLLECTION",
    "RENTAL_DEPT", "USE_STOCK", "SKU_LONG_DESC", "BLOCK_SALES",
    "BRAND COLLECTION", "LANGUAGE", "GENRE", "GEOGRAPHY",
    "ITEM_ATTRIB9", "ITEM_ATTRIB10", "ITEM_ATTRIB11", "ITEM_ATTRIB12",
    "ITEM_ATTRIB13", "ITEM_ATTRIB14", "ITEM_ATTRIB15",
    "ZERO_PRICE_VALID", "SEARCH_NAME", "ENABLE_UNIQUEID",
]

SKU_V2_FIELD_SPECS = [
    "NOT USE", "Character(1), M", "Character(16), M", "Character(60), M",
    "Numeric(20,2), O", "NOT USE", "NOT USE", "NOT USE", "NOT USE", "NOT USE",
    "Character(20),M", "NOT USE", "NOT USE", "NOT USE", "NOT USE",
    "Character(1),M", "NOT USE", "NOT USE", "NOT USE", "NOT USE",
    "NOT USE", "Character(1),M", "NOT USE", "NOT USE", "Character(1),M",
    "NOT USE", "NOT USE", "NOT USE",
    "Character(20),M", "Character(20),O", "Character(20),M", "Character(20),M",
    "Character(20),O", "Character(1),M", "Character(1000),O", "Character(1),M",
    "Character(20),O", "Character(20),O", "Character(20),O", "Character(20),O",
    "Character(20),O", "Character(20),O", "Character(20),O", "Character(20),O",
    "Character(20),O", "Character(20),O", "Character(20),O",
    "Char(1),O", "Char(20),O", "Char(1),O",
]

HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SPEC_FONT = Font(italic=True, size=8, color="666666")
THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))


def default_export_filename(today: date | None = None) -> str:
    run_date = today or date.today()
    return f"nec_jewel_master_data_{run_date.strftime('%Y%m%d')}.xlsx"


def _tenant_catg_code(form_factor: str) -> str:
    cag = CAG_CATG_MAP.get(form_factor, "DECORATIVE ITEM")
    return _CAG_TO_TENANT.get(cag, "VE_HD_DECOR")


def _style_header(ws: Any, row: int = 1) -> None:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws: Any, max_width: int = 30) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _sku_row(product: dict[str, Any]) -> list[Any]:
    sku_code = (product.get("sku_code") or "")[:16]
    desc = (product.get("description") or "")[:60]
    long_desc = (product.get("long_description") or product.get("description") or "")[:1000]
    form_factor = product.get("form_factor") or product.get("product_type") or ""
    attrs = product.get("attributes") or {}
    if isinstance(attrs, str):
        try:
            import json
            attrs = json.loads(attrs)
        except Exception:
            attrs = {}
    material = (attrs.get("materials") or attrs.get("material") or product.get("material") or "")[:20]

    return [
        None,
        "A",
        sku_code,
        desc,
        product.get("cost_price"),
        None,
        None,
        None,
        None,
        None,
        _tenant_catg_code(form_factor),
        None,
        None,
        None,
        None,
        TAX_CODE,
        None,
        None,
        None,
        None,
        None,
        "N",
        None,
        None,
        "Y",
        None,
        None,
        None,
        BRAND_NAME,
        (product.get("gender") or "UNISEX")[:20],
        "ADULT",
        "N.A",
        None,
        "Y" if product.get("use_stock", True) else "N",
        long_desc,
        "Y" if product.get("block_sales", False) else "N",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        material,
        None,
    ]


def build_catg_sheet(wb: openpyxl.Workbook, tenant_code: str = DEFAULT_TENANT_CODE) -> None:
    ws = wb.create_sheet("CATG")
    ws.append(["PARENT_CATG_CODE", "CHILD_CATG_CODE", "CATG_DESC", "CAG_CATG_CODE"])
    ws.append(["Character(20), M", "Character(20), M", "Character(60), M", "Character(20), M"])
    for parent, child, desc, cag in make_tenant_catg_tree(tenant_code):
        ws.append([parent, child, desc, cag or None])
    _style_header(ws)
    for col in range(1, 5):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws)


def build_sku_sheet(wb: openpyxl.Workbook, products: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("SKU v2")
    ws.append(SKU_V2_HEADERS)
    ws.append(SKU_V2_FIELD_SPECS)
    count = 0
    for product in products:
        if not product.get("sku_code"):
            continue
        ws.append(_sku_row(product))
        count += 1
    _style_header(ws)
    for col in range(1, len(SKU_V2_HEADERS) + 1):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws, max_width=25)
    return count


def build_plu_sheet(wb: openpyxl.Workbook, products: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("PLU")
    ws.append(["FILENAME", "MODE", "PLU_CODE", "SKU_CODE"])
    ws.append([None, "Character(1), M", "Character(20),M", "Character(16),M"])
    count = 0
    for product in products:
        plu = product.get("nec_plu") or product.get("plu_code")
        sku_code = product.get("sku_code")
        if not plu or not sku_code:
            continue
        ws.append([None, "A", str(plu), sku_code[:16]])
        count += 1
    _style_header(ws)
    for col in range(1, 5):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws)
    return count


def build_price_sheet(wb: openpyxl.Workbook, products: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("PRICE")
    ws.append(["MODE", "SKU_CODE", "STORE_ID", "PRICE_INCLTAX", "PRICE_EXCLTAX", "PRICE_UNIT", "PRICE_FRDATE", "PRICE_TODATE"])
    ws.append(["Character(1),M", "Char(16),M", "Char(5),O", "Numeric(20,2),M", "Numeric(20,2),M", "Numberic(8)", "Date(YYYYMMDD),M", "Date(YYYYMMDD),M"])
    count = 0
    today_str = date.today().strftime("%Y%m%d")
    for product in products:
        price_incl = product.get("retail_price") or product.get("price_incl_tax")
        sku_code = product.get("sku_code")
        if not price_incl or not sku_code:
            continue
        price_excl = product.get("price_excl_tax") or round(float(price_incl) / (1 + GST_RATE), 2)
        ws.append([
            "A",
            sku_code[:16],
            None,
            float(price_incl),
            float(price_excl),
            1,
            today_str,
            FAR_FUTURE,
        ])
        count += 1
    _style_header(ws)
    for col in range(1, 9):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws)
    return count


def build_promo_sheet(wb: openpyxl.Workbook, products: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("PROMO")
    ws.append(["DISC_ID", "TENANT_CATG_CODE", "SKU_CODE", "LINE_TYPE", "DISC_METHOD", "DISC_VALUE", "M&M_LINEGRP"])
    ws.append(["Character(20), M", "Character(20), O", "Character(16), O", "Character(10), M", "Character(20), M", "Numeric(11,2), M", "Character(1), O"])
    count = 0
    for disc_id, _label, pct in PROMO_TIERS:
        for product in products:
            sku_code = product.get("sku_code")
            if not sku_code:
                continue
            ws.append([disc_id, None, sku_code[:16], "Include", "PercentOff", pct, "A"])
            count += 1
    _style_header(ws)
    for col in range(1, 8):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws)
    return count


def build_invdetails_sheet(wb: openpyxl.Workbook, products: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("INVDETAILS")
    ws.append(["FILENAME", "SKU_CODE", "ACTION", "INV_VALUE"])
    ws.append([None, "Character(16),M", "Character(16),M", "Numeric(9)"])
    count = 0
    for product in products:
        qty = product.get("qty_on_hand")
        sku_code = product.get("sku_code")
        if not sku_code or qty is None:
            continue
        try:
            qty_int = int(qty)
        except (TypeError, ValueError):
            continue
        if qty_int <= 0:
            continue
        # Spec section 4.5: ACTION must be "Add", "Subtract" or "Update".
        ws.append([None, sku_code[:16], "Update", qty_int])
        count += 1
    _style_header(ws)
    for col in range(1, 5):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws)
    return count


def build_workbook(
    products: list[dict[str, Any]],
    tenant_code: str = DEFAULT_TENANT_CODE,
) -> tuple[openpyxl.Workbook, dict[str, int]]:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_catg_sheet(wb, tenant_code)
    counts = {
        "sku": build_sku_sheet(wb, products),
        "plu": build_plu_sheet(wb, products),
        "price": build_price_sheet(wb, products),
        "promo": build_promo_sheet(wb, products),
        "inventory": build_invdetails_sheet(wb, products),
    }
    return wb, counts


def build_workbook_bytes(
    products: list[dict[str, Any]],
    tenant_code: str = DEFAULT_TENANT_CODE,
) -> bytes:
    wb, _counts = build_workbook(products, tenant_code)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def fetch_sellable_skus_from_firestore(
    fs_db: Any,
    *,
    brand_name: str = BRAND_NAME,
    store_code: str | None = None,
    inv_store_code: str = DEFAULT_INV_STORE_CODE,
    include_drafts: bool = False,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Firestore equivalent of ``fetch_sellable_skus_from_session``.

    Returns the same ``(sellable, excluded)`` tuple shape so the workbook
    builders are unchanged. ``fs_db`` is a sync Firestore client (e.g.
    ``app.firestore.db``).
    """
    today_iso = date.today().isoformat()

    # Resolve inventory and (optional) target-store doc ids.
    inv_store_id: str | None = None
    target_store_id: str | None = None
    for s in fs_db.collection("stores").stream():
        sd = s.to_dict() or {}
        sc = sd.get("store_code")
        if sc == inv_store_code:
            inv_store_id = s.id
        if store_code and sc == store_code:
            target_store_id = s.id

    # Gather brand-matched SKUs across stores (or just one if store_code given).
    all_skus: list[dict[str, Any]] = []
    for s in fs_db.collection("stores").stream():
        if store_code and s.id != target_store_id:
            continue
        for sku_snap in s.reference.collection("inventory").stream():
            sku = sku_snap.to_dict() or {}
            if sku.get("brand_name") == brand_name:
                all_skus.append(sku)

    # PLU map (sku_id → plu_code) — top-level collection.
    plu_map: dict[str, str] = {}
    for p in fs_db.collection("plus").stream():
        pd = p.to_dict() or {}
        sid = pd.get("sku_id")
        plu = pd.get("plu_code")
        if sid and plu:
            plu_map[sid] = plu

    # Active-price map (sku_id → price doc) — scan every store's prices subcol.
    price_map: dict[str, dict[str, Any]] = {}
    for s in fs_db.collection("stores").stream():
        for pr in s.reference.collection("prices").stream():
            pd = pr.to_dict() or {}
            sid = pd.get("sku_id")
            if not sid:
                continue
            vf = pd.get("valid_from") or ""
            vt = pd.get("valid_to") or ""
            if vf <= today_iso <= vt:
                price_map[sid] = pd

    # Qty map: only from the inventory store.
    qty_map: dict[str, int] = {}
    if inv_store_id:
        for st in fs_db.collection(f"stores/{inv_store_id}/stock").stream():
            std = st.to_dict() or {}
            sid = std.get("sku_id")
            qoh = std.get("qty_on_hand")
            if sid is None or qoh is None:
                continue
            qty_map[sid] = qty_map.get(sid, 0) + int(qoh)

    sellable: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    for sku in all_skus:
        sku_id = sku.get("id")
        plu_code = plu_map.get(sku_id) if sku_id else None
        price = price_map.get(sku_id) if sku_id else None

        has_price = price is not None
        has_plu = plu_code is not None
        is_active = sku.get("status") == "active"
        not_blocked = not sku.get("block_sales")
        sale_ready = bool(sku.get("sale_ready"))
        has_desc = bool(sku.get("description"))

        full_sellable = (
            sale_ready and is_active and not_blocked
            and has_price and has_plu and has_desc
        )
        passes = (is_active and not_blocked) if include_drafts else full_sellable

        if passes:
            sellable.append({
                "sku_code": sku.get("sku_code"),
                "description": sku.get("description"),
                "long_description": sku.get("long_description"),
                "cost_price": sku.get("cost_price"),
                "product_type": sku.get("product_type"),
                "form_factor": sku.get("form_factor"),
                "attributes": sku.get("attributes"),
                "use_stock": sku.get("use_stock"),
                "block_sales": sku.get("block_sales"),
                "gender": sku.get("gender"),
                "age_group": sku.get("age_group"),
                "legacy_code": sku.get("legacy_code"),
                "sale_ready": sku.get("sale_ready"),
                "status": sku.get("status"),
                "retail_price": price.get("price_incl_tax") if price else None,
                "price_excl_tax": price.get("price_excl_tax") if price else None,
                "nec_plu": plu_code,
                "qty_on_hand": qty_map.get(sku_id, 0) if sku_id else 0,
            })

        if not full_sellable:
            excluded.append({
                "sku_code": sku.get("sku_code"),
                "description": sku.get("description"),
                "form_factor": sku.get("form_factor"),
                "sale_ready": sku.get("sale_ready"),
                "status": sku.get("status"),
                "block_sales": sku.get("block_sales"),
                "has_price": has_price,
                "has_plu": has_plu,
            })

    sellable.sort(key=lambda r: r.get("sku_code") or "")
    excluded.sort(key=lambda r: r.get("sku_code") or "")
    return sellable, excluded
