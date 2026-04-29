"""Service for importing legacy timesheet data from CSV or Excel files."""
from __future__ import annotations

import csv
import io
import uuid as _uuid
from datetime import datetime, date, time, timezone
from typing import Any, Dict, Optional
from uuid import UUID

from openpyxl import load_workbook
from google.cloud.firestore_v1.client import Client as FirestoreClient

import app.firestore as _fs  # access as `_fs.db` so init stays lazy
from app.firestore_helpers import query_collection


class ImportError:
    """A single row-level import error."""

    def __init__(self, row: int, message: str):
        self.row = row
        self.message = message


class ImportResult:
    """Aggregate result of an import operation."""

    def __init__(self):
        self.imported_count = 0
        self.skipped_count = 0
        self.errors: list[ImportError] = []

    def to_dict(self) -> dict:
        return {
            "imported_count": self.imported_count,
            "skipped_count": self.skipped_count,
            "errors": [{"row": e.row, "message": e.message} for e in self.errors],
        }


# Column name normalisation mapping
_COLUMN_ALIASES = {
    "staff_name": "staff_name",
    "name": "staff_name",
    "employee_name": "staff_name",
    "employee": "staff_name",
    "email": "email",
    "staff_email": "email",
    "date": "date",
    "clock_in_time": "clock_in_time",
    "clock_in": "clock_in_time",
    "start_time": "clock_in_time",
    "clock_out_time": "clock_out_time",
    "clock_out": "clock_out_time",
    "end_time": "clock_out_time",
    "break_minutes": "break_minutes",
    "break": "break_minutes",
    "notes": "notes",
    "note": "notes",
    "comments": "notes",
}


def _normalise_header(header: str) -> str | None:
    """Map a raw header string to a canonical column name."""
    return _COLUMN_ALIASES.get(header.strip().lower().replace(" ", "_"))


def _parse_rows_from_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        return []
    # Build normalised mapping
    col_map = {}
    for raw in reader.fieldnames:
        norm = _normalise_header(raw)
        if norm:
            col_map[raw] = norm
    rows = []
    for raw_row in reader:
        row = {col_map[k]: v for k, v in raw_row.items() if k in col_map}
        rows.append(row)
    return rows


def _parse_rows_from_excel(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            norm = _normalise_header(str(cell))
            if norm:
                col_map[idx] = norm
    rows = []
    for data_row in rows_iter:
        row: dict[str, str] = {}
        for idx, norm_name in col_map.items():
            val = data_row[idx] if idx < len(data_row) else None
            row[norm_name] = str(val) if val is not None else ""
        rows.append(row)
    wb.close()
    return rows


def _parse_date(value: str) -> date | None:
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return None


def _parse_time(value: str) -> time | None:
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
        try:
            return datetime.strptime(value.strip(), fmt).time()
        except (ValueError, AttributeError):
            continue
    return None


def _build_user_lookup(db: FirestoreClient) -> dict[str, dict]:
    """Build lookup dicts keyed by lowercase email and full_name."""
    users = sorted(
        query_collection("users"),
        key=lambda user: (
            str(user.get("updated_at") or ""),
            str(user.get("created_at") or ""),
            str(user.get("id") or ""),
        ),
    )
    lookup: dict[str, dict] = {}
    for u in users:
        if u.get("email"):
            lookup[u["email"].lower()] = u
        if u.get("full_name"):
            lookup[u["full_name"].lower()] = u
    return lookup


def _entry_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).date()
        except ValueError:
            return None
    return None


def import_timesheet_file(
    db: FirestoreClient,
    store_id: UUID,
    filename: str,
    content: bytes,
) -> ImportResult:
    """Parse and import a CSV or Excel timesheet file."""
    result = ImportResult()

    lower_name = filename.lower()
    if lower_name.endswith((".xlsx", ".xls")):
        rows = _parse_rows_from_excel(content)
    elif lower_name.endswith(".csv"):
        rows = _parse_rows_from_csv(content)
    else:
        result.errors.append(ImportError(0, "Unsupported file format. Use CSV or Excel (.xlsx)."))
        return result

    if not rows:
        result.errors.append(ImportError(0, "File is empty or has no data rows."))
        return result

    user_lookup = _build_user_lookup(db)
    col_path = f"stores/{store_id}/timesheets"
    existing_entry_keys = {
        (str(entry.get("user_id", "")), entry_date)
        for entry in query_collection(col_path, order_by="-clock_in")
        if (entry_date := _entry_date(entry.get("clock_in"))) is not None
    }

    # Use batch writes for bulk creation
    batch = _fs.db.batch()
    batch_count = 0

    for row_idx, row in enumerate(rows, start=2):  # row 1 is header
        # --- Resolve user ---
        email_val = row.get("email", "").strip()
        name_val = row.get("staff_name", "").strip()
        user: dict | None = None
        if email_val:
            user = user_lookup.get(email_val.lower())
        if user is None and name_val:
            user = user_lookup.get(name_val.lower())
        if user is None:
            identifier = email_val or name_val or "(empty)"
            result.errors.append(ImportError(row_idx, f"Staff not found: {identifier}"))
            result.skipped_count += 1
            continue

        # --- Parse date ---
        date_val = row.get("date", "").strip()
        entry_date = _parse_date(date_val)
        if entry_date is None:
            result.errors.append(ImportError(row_idx, f"Invalid date: {date_val}"))
            result.skipped_count += 1
            continue

        # --- Parse times ---
        clock_in_str = row.get("clock_in_time", "").strip()
        clock_out_str = row.get("clock_out_time", "").strip()
        clock_in_time = _parse_time(clock_in_str)
        clock_out_time = _parse_time(clock_out_str)
        if clock_in_time is None:
            result.errors.append(ImportError(row_idx, f"Invalid clock-in time: {clock_in_str}"))
            result.skipped_count += 1
            continue
        if clock_out_time is None:
            result.errors.append(ImportError(row_idx, f"Invalid clock-out time: {clock_out_str}"))
            result.skipped_count += 1
            continue

        clock_in_dt = datetime.combine(entry_date, clock_in_time)
        clock_out_dt = datetime.combine(entry_date, clock_out_time)

        if clock_out_dt <= clock_in_dt:
            result.errors.append(ImportError(row_idx, "Clock-out must be after clock-in"))
            result.skipped_count += 1
            continue

        # --- Parse break_minutes ---
        break_str = row.get("break_minutes", "0").strip()
        try:
            break_minutes = int(break_str) if break_str else 0
        except ValueError:
            result.errors.append(ImportError(row_idx, f"Invalid break minutes: {break_str}"))
            result.skipped_count += 1
            continue

        user_id = user.get("id", "")
        user_name = user.get("full_name", "Unknown")

        duplicate_key = (str(user_id), entry_date)
        if duplicate_key in existing_entry_keys:
            result.errors.append(
                ImportError(row_idx, f"Duplicate entry for {user_name} on {entry_date}")
            )
            result.skipped_count += 1
            continue

        # --- Create TimeEntry via batch ---
        notes = row.get("notes", "").strip() or None
        doc_id = str(_uuid.uuid4())
        now = datetime.now(timezone.utc)
        ref = _fs.db.collection(col_path).document(doc_id)
        batch.set(ref, {
            "id": doc_id,
            "user_id": str(user_id),
            "store_id": str(store_id),
            "clock_in": clock_in_dt,
            "clock_out": clock_out_dt,
            "break_minutes": break_minutes,
            "notes": notes,
            "status": "approved",
            "approved_by": None,
            "user_name": user_name,
            "created_at": now,
            "updated_at": now,
        })
        batch_count += 1
        result.imported_count += 1
        existing_entry_keys.add(duplicate_key)

        # Firestore batch limit is 500
        if batch_count >= 499:
            batch.commit()
            batch = _fs.db.batch()
            batch_count = 0

    if batch_count > 0:
        batch.commit()
    return result
