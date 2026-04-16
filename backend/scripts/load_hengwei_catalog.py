"""Load Hengwei Craft full product catalog from Excel files.

Parses both catalogs:
  - 衡威.爱达荷目录2026(英).xlsx (Idaho 2026): 摆件, 花瓶, 托盘+盒子, 挂画
  - 衡威家居目录2026(英).xlsx (Home 2026): 铜水晶系列1-5, 挂饰系列6, 可订做产品

Products are in a grid layout (4 per row block, 7 rows per block).
Creates SKUs, inventory, and SupplierProduct links.

Usage:
    cd backend
    DATABASE_URL="..." python -m scripts.load_hengwei_catalog
"""
from __future__ import annotations

import asyncio
import os
import re
import sys
import uuid
from datetime import datetime
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models.store import Store
from app.models.supplier import Supplier, SupplierProduct
from app.models.inventory import SKU, Category, Brand, Inventory

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql+asyncpg://retailsg:RetailSG2026Secure@127.0.0.1:5434/retailsg",
)

STORE_CODE = "JEWEL-01"
SUPPLIER_CODE = "CN-001"
CNY_TO_SGD = Decimal("5.34")

CATALOG_DIR = Path(os.environ.get(
    "CATALOG_DIR",
    str(Path(__file__).resolve().parents[2] / "docs" / "suppliers" / "hengweicraft" / "catalog"),
))

# Category mapping from sheet names
SHEET_CATEGORIES = {
    "摆件": "Decorations (摆件)",
    "花瓶": "Vases (花瓶)",
    "托盘+盒子": "Trays & Boxes (托盘+盒子)",
    "挂画": "Wall Art (挂画)",
    "铜、水晶系列1": "Copper & Crystal Series 1",
    "铜、水晶系列2": "Copper & Crystal Series 2",
    "铜、水晶系列3": "Copper & Crystal Series 3",
    "铜、水晶系列4": "Copper & Crystal Series 4",
    "铜、水晶系列5": "Copper & Crystal Series 5",
    "挂饰系列6": "Hanging Decorations (挂饰)",
    "可订做产品": "Custom Products (可订做)",
}


# ------------------------------------------------------------------ #
# Excel parser                                                          #
# ------------------------------------------------------------------ #

def parse_price(val) -> Decimal | None:
    if val is None:
        return None
    s = str(val).strip()
    nums = re.findall(r"[\d.]+", s)
    return Decimal(nums[0]) if nums else None


def parse_catalog_sheet(ws, sheet_name: str, catalog_name: str) -> list[dict]:
    col_pairs = [(1, 2), (4, 5), (7, 8), (10, 11)]
    products = []

    row = 1
    while row <= ws.max_row:
        cell = ws.cell(row=row, column=1).value
        if cell and str(cell).strip().lower() == "model":
            for label_col, value_col in col_pairs:
                model = ws.cell(row=row, column=value_col).value
                if not model:
                    continue
                model = str(model).strip()
                if not model or model.lower() == "model":
                    continue

                name = str(ws.cell(row=row + 1, column=value_col).value or "").strip()
                size = str(ws.cell(row=row + 2, column=value_col).value or "").strip()
                materials = str(ws.cell(row=row + 3, column=value_col).value or "").strip()
                color = str(ws.cell(row=row + 4, column=value_col).value or "").strip()

                # Price row
                price_val = ws.cell(row=row + 5, column=value_col).value
                if price_val is None:
                    label = ws.cell(row=row + 5, column=label_col).value
                    if label and str(label).strip().lower() in ("unit price", "special"):
                        price_val = ws.cell(row=row + 5, column=value_col).value

                price = parse_price(price_val)

                # Extract primary code (first token), truncate to 16 chars for DB
                code = model.split()[0].strip().rstrip(",;")[:16]

                products.append({
                    "code": code,
                    "full_model": model,
                    "name": name or "decoration",
                    "size": size,
                    "materials": materials,
                    "color": color,
                    "price_cny": price,
                    "sheet": sheet_name,
                    "catalog": catalog_name,
                })
            row += 7
        else:
            row += 1

    return products


def load_all_catalogs() -> dict[str, dict]:
    """Parse both Excel catalogs and return deduplicated products keyed by code."""
    all_products: list[dict] = []

    f1 = CATALOG_DIR / "衡威.爱达荷目录2026(英).xlsx"
    if f1.exists():
        wb = openpyxl.load_workbook(str(f1), data_only=True)
        for name in wb.sheetnames:
            prods = parse_catalog_sheet(wb[name], name, "Idaho 2026")
            all_products.extend(prods)
            print(f"  Idaho/{name}: {len(prods)} products")

    f2 = CATALOG_DIR / "衡威家居目录2026(英).xlsx"
    if f2.exists():
        wb = openpyxl.load_workbook(str(f2), data_only=True)
        for name in wb.sheetnames:
            prods = parse_catalog_sheet(wb[name], name, "Home 2026")
            all_products.extend(prods)
            print(f"  Home/{name}: {len(prods)} products")

    # Deduplicate — keep first occurrence (preserves catalog ordering)
    seen: dict[str, dict] = {}
    for p in all_products:
        if p["code"] not in seen:
            seen[p["code"]] = p

    print(f"\n  Total parsed: {len(all_products)}, Unique codes: {len(seen)}")
    return seen


# ------------------------------------------------------------------ #
# Database loader                                                       #
# ------------------------------------------------------------------ #

async def get_or_create_category(
    session: AsyncSession, store_id: uuid.UUID, code: str, description: str
) -> Category:
    result = await session.execute(
        select(Category).where(Category.catg_code == code, Category.store_id == store_id)
    )
    cat = result.scalar_one_or_none()
    if not cat:
        cat = Category(
            id=uuid.uuid4(), store_id=store_id, catg_code=code, description=description
        )
        session.add(cat)
        await session.flush()
    return cat


async def main():
    connect_args: dict = {}
    if "127.0.0.1" in DATABASE_URL or "localhost" in DATABASE_URL:
        connect_args["ssl"] = False

    engine = create_async_engine(DATABASE_URL, connect_args=connect_args)
    async_session = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    # Parse catalogs
    print("=== Parsing Hengwei Craft catalogs ===\n")
    catalog = load_all_catalogs()

    async with async_session() as session:
        # Look up store, supplier, brand
        result = await session.execute(select(Store).where(Store.store_code == STORE_CODE))
        store = result.scalar_one_or_none()
        if not store:
            print(f"ERROR: Store {STORE_CODE} not found.")
            return

        result = await session.execute(select(Supplier).where(Supplier.supplier_code == SUPPLIER_CODE))
        supplier = result.scalar_one_or_none()
        if not supplier:
            print(f"ERROR: Supplier {SUPPLIER_CODE} not found.")
            return

        # Get or create Hengwei brand
        result = await session.execute(select(Brand).where(Brand.name == "Hengwei Craft"))
        brand = result.scalar_one_or_none()
        if not brand:
            brand = Brand(id=uuid.uuid4(), name="Hengwei Craft", category_type="Decorative Arts")
            session.add(brand)
            await session.flush()

        # Create categories for each sheet
        cat_map: dict[str, Category] = {}
        for sheet_name, desc in SHEET_CATEGORIES.items():
            code = "HW-" + re.sub(r"[^A-Z0-9]", "", desc.upper())[:15]
            cat = await get_or_create_category(session, store.id, code, f"Hengwei — {desc}")
            cat_map[sheet_name] = cat

        # Also keep the generic DECO-ARTS as fallback
        fallback_cat = await get_or_create_category(
            session, store.id, "DECO-ARTS",
            "Decorative Arts — Copper, Crystal, Marble, Malachite"
        )

        # Load products
        created = 0
        updated = 0
        print(f"\n=== Loading {len(catalog)} products ===\n")

        for code, item in sorted(catalog.items()):
            category = cat_map.get(item["sheet"], fallback_cat)
            price_cny = item["price_cny"] or Decimal("0")
            cost_sgd = (price_cny / CNY_TO_SGD).quantize(Decimal("0.01")) if price_cny else Decimal("0")

            desc_short = item["name"] or "decoration"
            if len(desc_short) < 5 and item["materials"]:
                desc_short = f"{desc_short} — {item['materials']}"
            desc_short = desc_short[:60]

            long_desc = f"{item['name']}. Materials: {item['materials']}. Size: {item['size']}."
            if item["color"]:
                long_desc += f" Color: {item['color']}."
            long_desc += f" Model: {item['full_model']}. Catalog: {item['catalog']}/{item['sheet']}."

            # Upsert SKU
            result = await session.execute(
                select(SKU).where(SKU.sku_code == code, SKU.store_id == store.id)
            )
            sku = result.scalar_one_or_none()

            if sku:
                sku.description = desc_short
                sku.long_description = long_desc[:1000]
                sku.cost_price = cost_sgd
                sku.category_id = category.id
                sku.brand_id = brand.id
                updated += 1
            else:
                sku = SKU(
                    id=uuid.uuid4(),
                    store_id=store.id,
                    sku_code=code,
                    description=desc_short,
                    long_description=long_desc[:1000],
                    cost_price=cost_sgd,
                    category_id=category.id,
                    brand_id=brand.id,
                    tax_code="E",
                    is_unique_piece=price_cny >= 400 if price_cny else False,
                )
                session.add(sku)
                created += 1

            await session.flush()

            # Ensure inventory
            result = await session.execute(
                select(Inventory).where(Inventory.sku_id == sku.id, Inventory.store_id == store.id)
            )
            if not result.scalar_one_or_none():
                session.add(Inventory(
                    id=uuid.uuid4(),
                    sku_id=sku.id,
                    store_id=store.id,
                    qty_on_hand=0,  # catalog item, not yet purchased
                    reorder_level=0,
                    reorder_qty=1,
                    last_updated=datetime.utcnow(),
                ))

            # Link to supplier
            result = await session.execute(
                select(SupplierProduct).where(
                    SupplierProduct.supplier_id == supplier.id,
                    SupplierProduct.sku_id == sku.id,
                )
            )
            if not result.scalar_one_or_none():
                session.add(SupplierProduct(
                    id=uuid.uuid4(),
                    supplier_id=supplier.id,
                    sku_id=sku.id,
                    supplier_sku_code=code,
                    supplier_unit_cost=price_cny or Decimal("0"),
                    currency="CNY",
                    min_order_qty=1,
                    lead_time_days=30,
                    is_preferred=True,
                ))

        await session.commit()
        print(f"Done. Created: {created}, Updated: {updated}")
        print(f"Total Hengwei products in database: {created + updated}")

    await engine.dispose()


if __name__ == "__main__":
    asyncio.run(main())
