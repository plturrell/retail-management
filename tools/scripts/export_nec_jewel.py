#!/usr/bin/env python3
"""
Export sellable catalogue → Jewel NEC POS Master Data Excel workbook.

Pulls from the live RetailSG database (Cloud SQL) by default so only
products that are truly ready for sale reach the NEC POS system. A SKU
is *sellable* when ALL of:
  * ``sale_ready = true``
  * ``status = 'active'`` and ``block_sales = false``
  * an active ``prices`` row exists (valid_from <= today <= valid_to)
  * a description is present
  * a PLU (NEC barcode) row exists

Everything else is explicitly excluded and reported as a gap so the operator
can complete the missing data before re-running.

Output sheets: CATG, SKU v2, PLU, PRICE, PROMO, INVDETAILS

Usage:
  # Default -- read from Cloud SQL (requires cloud-sql-proxy on :5434)
  python export_nec_jewel.py

  # Legacy -- read from the JSON master list (no DB required)
  python export_nec_jewel.py --from-json

  # Explicit DB URL
  DATABASE_URL="postgresql+asyncpg://..." python export_nec_jewel.py

Filters (DB mode):
  --brand "VICTORIA ENSO"    only products for this brand (default)
  --store JEWEL-01           only SKUs stocked at this store
  --include-drafts           bypass the sale_ready gate (debug only)
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
import time
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = "data/master_product_list.json"
DEFAULT_OUTPUT = "data/exports/nec_jewel_master_data.xlsx"
DEFAULT_DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql+asyncpg://retailsg:RetailSG2026Secure@127.0.0.1:5434/retailsg",
)

TENANT_CODE = "[TENANT]"  # Jewel assigns the real code after registration
BRAND_NAME = "VICTORIA ENSO"
TAX_CODE = "G"  # GST-taxable
GST_RATE = 0.09  # Singapore 9% GST
TODAY_STR = date.today().strftime("%Y%m%d")
FAR_FUTURE = "20991231"


# ── Promotion Tiers ────────────────────────────────────────────────────────
# Each tuple: (DISC_ID, label, percent_off)

PROMO_TIERS: list[tuple[str, str, float]] = [
    ("VE_GENERAL_10",     "10% General",              10),
    ("VE_SPECIAL_15",     "15% Special",              15),
    ("VE_DIRECTOR_20",    "20% Director Approved",    20),
    ("VE_CLEARANCE_25",   "25% Clearance Special",    25),
    ("VE_STAFFCARD_20",   "20% Jewel Staff Card",     20),
    ("VE_VIP_LOYALTY_20", "20% VIP Loyalty Customer", 20),
]


# ── CAG Category Hierarchy Mapping ───────────────────────────────────────────
# Maps form_factor (or legacy product_type) -> Jewel CAG Level 3 category code

CAG_CATG_MAP: dict[str, str] = {
    # Jewellery
    "Bracelet":         "BANGLE & BRACELETS",
    "Necklace":         "NECKLACE",
    "Ring":             "RINGS",
    "Earring":          "EARRINGS",
    "Charm":            "CHARMS",
    "Pendant":          "CHARMS",
    "Bead Strand":      "COSTUME JEWELLERY",
    "Accessory":        "JEWELLERY ACCESSORY",
    # Specimen Minerals
    "Loose Gemstone":   "PRECIOUS STONE/GOLD",
    "Raw Specimen":     "PRECIOUS STONE/GOLD",
    "Crystal Cluster":  "PRECIOUS STONE/GOLD",
    "Crystal Point":    "PRECIOUS STONE/GOLD",
    "Tumbled Stone":    "PRECIOUS STONE/GOLD",
    "Gemstone Bead":    "PRECIOUS STONE/GOLD",
    "Healing Crystal":  "PRECIOUS STONE/GOLD",
    # Home Decor
    "Figurine":          "DECORATIVE ITEM",
    "Sculpture":         "DECORATIVE ITEM",
    "Bookend":           "DECORATIVE ITEM",
    "Bowl":              "DECORATIVE ITEM",
    "Vase":              "DECORATIVE ITEM",
    "Box":               "DECORATIVE ITEM",
    "Tray":              "DECORATIVE ITEM",
    "Decorative Object": "DECORATIVE ITEM",
    "Wall Art":          "DECORATIVE ITEM",
    "Gift Set":          "GENERAL SOUVENIRS",
    "Repair Service":    "REPAIR SERVICE",
}

# Three top-level tenant categories: JEWELLERY, HOME DECOR, SPECIMEN MINERALS
# (PARENT_CATG_CODE, CHILD_CATG_CODE, CATG_DESC, CAG_CATG_CODE)
TENANT_CATG_TREE: list[tuple[str, str, str, str]] = [
    # Level 1
    (TENANT_CODE,    "VE_JEWELLERY",   "Jewellery",                   ""),
    (TENANT_CODE,    "VE_HOMEDECOR",   "Home Decor",                  ""),
    (TENANT_CODE,    "VE_MINERALS",    "Specimen Minerals",           ""),
    # Level 2 -- Jewellery
    ("VE_JEWELLERY", "VE_JW_BRACELET", "Bracelets",                   "BANGLE & BRACELETS"),
    ("VE_JEWELLERY", "VE_JW_NECKLACE", "Necklaces",                   "NECKLACE"),
    ("VE_JEWELLERY", "VE_JW_RING",     "Rings",                       "RINGS"),
    ("VE_JEWELLERY", "VE_JW_EARRING",  "Earrings",                    "EARRINGS"),
    ("VE_JEWELLERY", "VE_JW_CHARM",    "Charms & Pendants",           "CHARMS"),
    ("VE_JEWELLERY", "VE_JW_COSTUME",  "Costume Jewellery",           "COSTUME JEWELLERY"),
    ("VE_JEWELLERY", "VE_JW_ACC",      "Jewellery Accessory",         "JEWELLERY ACCESSORY"),
    ("VE_JEWELLERY", "VE_JW_REPAIR",   "Repair Service",              "REPAIR SERVICE"),
    # Level 2 -- Home Decor
    ("VE_HOMEDECOR", "VE_HD_DECOR",    "Decorative Items",            "DECORATIVE ITEM"),
    ("VE_HOMEDECOR", "VE_HD_GENERAL",  "Gifts & General",             "GENERAL SOUVENIRS"),
    # Level 2 -- Specimen Minerals
    ("VE_MINERALS",  "VE_SM_STONE",    "Precious Stones & Crystals",  "PRECIOUS STONE/GOLD"),
]

_CAG_TO_TENANT: dict[str, str] = {
    cag: child for _p, child, _d, cag in TENANT_CATG_TREE if cag
}


def _tenant_catg_code(form_factor: str) -> str:
    """Map form_factor -> tenant category child code for SKU_CATG_TENANT."""
    cag = CAG_CATG_MAP.get(form_factor, "DECORATIVE ITEM")
    return _CAG_TO_TENANT.get(cag, "VE_HD_DECOR")


# ── SKU v2 Column Spec ───────────────────────────────────────────────────────

SKU_V2_HEADERS = [
    "FILENAME", "MODE", "SKU_CODE", "SKU_DESC", "COSTPRICE", "AVGCOST",
    "SKU_PRICE", "SKU_PRICE1", "SKU-EDATE1", "SKU-ETIME1",
    "SKU_CATG_TENANT", "SKU_CATG_CAG", "SKU_ROL", "SKU_ROQ", "SKU_QOH",
    "TAX_CODE", "PRO-FRDATE", "PRO_FRTIME", "PRO_TODATE", "PRO_TOTIME",
    "PRO_PRICE", "OPEN_ITEM", "KIT", "REMARKS", "SKU_DISC",
    "SKU_PRICE_LALO", "SKU_PRICE_HALO", "MIN_AGE",
    "BRAND", "GENDER", "AGE GROUP", "CHANGI COLLECTION",
    "RENTAL_DEPT", "USE_STOCK", "SKU_LONG_DESC", "BLOCK_SALES",
    "BRAND COLLECTION", "LANGUAGE", "GENRE", "GEOGRAPHY",
    "ITEM_ATTRIB9", "ITEM_ATTRIB10", "ITEM_ATTRIB11", "ITEM_ATTRIB12",
    "ITEM_ATTRIB13", "ITEM_ATTRIB14", "ITEM_ATTRIB15",
    "ZERO_PRICE_VALID", "SEARCH_NAME", "ENABLE_UNIQUEID",
]

SKU_V2_FIELD_SPECS = [
    "NOT USE", "Character(1), M", "Character(16), M", "Character(60), M",
    "Numeric(20,2), O", "NOT USE", "NOT USE", "NOT USE", "NOT USE", "NOT USE",
    "Character(20),M", "NOT USE", "NOT USE", "NOT USE", "NOT USE",
    "Character(1),M", "NOT USE", "NOT USE", "NOT USE", "NOT USE",
    "NOT USE", "Character(1),M", "NOT USE", "NOT USE", "Character(1),M",
    "NOT USE", "NOT USE", "NOT USE",
    "Character(20),M", "Character(20),O", "Character(20),M", "Character(20),M",
    "Character(20),O", "Character(1),M", "Character(1000),O", "Character(1),M",
    "Character(20),O", "Character(20),O", "Character(20),O", "Character(20),O",
    "Character(20),O", "Character(20),O", "Character(20),O", "Character(20),O",
    "Character(20),O", "Character(20),O", "Character(20),O",
    "Char(1),O", "Char(20),O", "Char(1),O",
]


def _sku_row(p: dict[str, Any]) -> list[Any]:
    """Build one SKU v2 row from a product dict (DB or JSON shape)."""
    # NEC SKU_CODE field is Character(16). Our codes are 14 chars max (XXX-XXX-000000).
    sku_code = (p.get("sku_code") or "")[:16]
    desc = (p.get("description") or "")[:60]
    long_desc = (p.get("long_description") or p.get("description") or "")[:1000]
    cost = p.get("cost_price")
    # form_factor drives the category mapping
    form_factor = p.get("form_factor") or p.get("product_type") or ""
    catg_tenant = _tenant_catg_code(form_factor)
    use_stock = "Y" if p.get("use_stock", True) else "N"
    block_sales = "Y" if p.get("block_sales", False) else "N"
    gender = (p.get("gender") or "UNISEX")[:20]
    # SEARCH_NAME: pull material from attributes JSONB or top-level key
    attrs = p.get("attributes") or {}
    if isinstance(attrs, str):
        try:
            attrs = json.loads(attrs)
        except Exception:
            attrs = {}
    material = (attrs.get("materials") or attrs.get("material") or p.get("material") or "")[:20]

    return [
        None,        # FILENAME (not used)
        "A",         # MODE -- Add
        sku_code,    # SKU_CODE
        desc,        # SKU_DESC
        cost,        # COSTPRICE
        None,        # AVGCOST (not used)
        None,        # SKU_PRICE (not used)
        None,        # SKU_PRICE1 (not used)
        None,        # SKU-EDATE1 (not used)
        None,        # SKU-ETIME1 (not used)
        catg_tenant, # SKU_CATG_TENANT
        None,        # SKU_CATG_CAG (not used)
        None,        # SKU_ROL (not used)
        None,        # SKU_ROQ (not used)
        None,        # SKU_QOH (not used)
        TAX_CODE,    # TAX_CODE
        None,        # PRO-FRDATE (not used)
        None,        # PRO_FRTIME (not used)
        None,        # PRO_TODATE (not used)
        None,        # PRO_TOTIME (not used)
        None,        # PRO_PRICE (not used)
        "N",         # OPEN_ITEM
        None,        # KIT (not used)
        None,        # REMARKS
        "Y",         # SKU_DISC -- discountable
        None,        # SKU_PRICE_LALO (not used)
        None,        # SKU_PRICE_HALO (not used)
        None,        # MIN_AGE
        BRAND_NAME,  # BRAND
        gender,      # GENDER
        "ADULT",     # AGE GROUP
        "N.A",       # CHANGI COLLECTION
        None,        # RENTAL_DEPT
        use_stock,   # USE_STOCK
        long_desc,   # SKU_LONG_DESC
        block_sales, # BLOCK_SALES
        None,        # BRAND COLLECTION
        None,        # LANGUAGE
        None,        # GENRE
        None,        # GEOGRAPHY
        None,        # ITEM_ATTRIB9
        None,        # ITEM_ATTRIB10
        None,        # ITEM_ATTRIB11
        None,        # ITEM_ATTRIB12
        None,        # ITEM_ATTRIB13
        None,        # ITEM_ATTRIB14
        None,        # ITEM_ATTRIB15
        None,        # ZERO_PRICE_VALID
        material,    # SEARCH_NAME
        None,        # ENABLE_UNIQUEID
    ]


# ── Styling helpers ──────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SPEC_FONT = Font(italic=True, size=8, color="666666")
THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))


def _style_header(ws: Any, row: int = 1) -> None:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws: Any, max_width: int = 30) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_catg_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("CATG")
    headers = ["PARENT_CATG_CODE", "CHILD_CATG_CODE", "CATG_DESC", "CAG_CATG_CODE"]
    specs = ["Character(20), M", "Character(20), M", "Character(60), M", "Character(20), M"]
    ws.append(headers)
    ws.append(specs)
    for parent, child, desc, cag in TENANT_CATG_TREE:
        ws.append([parent, child, desc, cag or None])
    _style_header(ws)
    for col in range(1, 5):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws)


def build_sku_sheet(wb: openpyxl.Workbook, products: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("SKU v2")
    ws.append(SKU_V2_HEADERS)
    ws.append(SKU_V2_FIELD_SPECS)
    count = 0
    for p in products:
        if not p.get("sku_code"):
            continue
        ws.append(_sku_row(p))
        count += 1
    _style_header(ws)
    for col in range(1, len(SKU_V2_HEADERS) + 1):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws, max_width=25)
    return count


def build_plu_sheet(wb: openpyxl.Workbook, products: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("PLU")
    headers = ["FILENAME", "MODE", "PLU_CODE", "SKU_CODE"]
    specs = [None, "Character(1), M", "Character(20),M", "Character(16),M"]
    ws.append(headers)
    ws.append(specs)
    count = 0
    for p in products:
        plu = p.get("nec_plu") or p.get("plu_code")
        if not plu or not p.get("sku_code"):
            continue
        ws.append([None, "A", str(plu), (p["sku_code"])[:16]])
        count += 1
    _style_header(ws)
    for col in range(1, 5):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws)
    return count


def build_price_sheet(wb: openpyxl.Workbook, products: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("PRICE")
    headers = ["MODE", "SKU_CODE", "STORE_ID", "PRICE_INCLTAX", "PRICE_EXCLTAX",
               "PRICE_UNIT", "PRICE_FRDATE", "PRICE_TODATE"]
    specs = ["Character(1),M", "Char(16),M", "Char(5),O", "Numeric(20,2),M",
             "Numeric(20,2),M", "Numberic(8)", "Date(YYYYMMDD),M", "Date(YYYYMMDD),M"]
    ws.append(headers)
    ws.append(specs)
    count = 0
    for p in products:
        price_incl = p.get("retail_price") or p.get("price_incl_tax")
        if not price_incl or not p.get("sku_code"):
            continue
        price_excl = p.get("price_excl_tax") or round(float(price_incl) / (1 + GST_RATE), 2)
        ws.append([
            "A",
            (p["sku_code"])[:16],
            None,
            float(price_incl),
            float(price_excl),
            1,
            TODAY_STR,
            FAR_FUTURE,
        ])
        count += 1
    _style_header(ws)
    for col in range(1, 9):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws)
    return count


def build_promo_sheet(wb: openpyxl.Workbook, products: list[dict[str, Any]]) -> int:
    """Build PROMO sheet with the 6 Victoria Enso discount tiers, applied per-SKU."""
    ws = wb.create_sheet("PROMO")
    headers = ["DISC_ID", "TENANT_CATG_CODE", "SKU_CODE", "LINE_TYPE",
               "DISC_METHOD", "DISC_VALUE", "M&M_LINEGRP"]
    specs = ["Character(20), M", "Character(20), O", "Character(16), O",
             "Character(10), M", "Character(20), M", "Numeric(11,2), M", "Character(1), O"]
    ws.append(headers)
    ws.append(specs)
    count = 0
    for disc_id, _label, pct in PROMO_TIERS:
        for p in products:
            sku = p.get("sku_code")
            if not sku:
                continue
            ws.append([disc_id, None, sku[:16], "Include", "PercentOff", pct, "A"])
            count += 1
    _style_header(ws)
    for col in range(1, 8):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws)
    return count


def build_invdetails_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("INVDETAILS")
    headers = ["FILENAME", "SKU_CODE", "ACTION", "INV_VALUE"]
    specs = [None, "Character(16),M", "Character(16),M", "Numeric(9)"]
    ws.append(headers)
    ws.append(specs)
    _style_header(ws)
    for col in range(1, 5):
        ws.cell(row=2, column=col).font = SPEC_FONT
    _auto_width(ws)


# ── DB fetch ──────────────────────────────────────────────────────────────────

async def fetch_sellable_skus(
    database_url: str,
    brand_name: str = BRAND_NAME,
    store_code: str | None = None,
    include_drafts: bool = False,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Query DB for sellable SKUs. Returns (sellable, excluded) lists."""
    try:
        from sqlalchemy.ext.asyncio import create_async_engine
        from sqlalchemy import text
    except ImportError:
        print("ERROR: sqlalchemy[asyncio] not installed. Run: pip install sqlalchemy asyncpg")
        sys.exit(1)

    engine = create_async_engine(database_url, echo=False)

    # Build WHERE clauses conditionally to avoid asyncpg NULL/boolean type inference issues
    sellability_clause = "" if include_drafts else """
          AND s.sale_ready = true
          AND pr.id IS NOT NULL
          AND pl.id IS NOT NULL
          AND s.description IS NOT NULL
          AND s.description != ''"""

    store_clause = "" if not store_code else "          AND st.store_code = :store_code"

    sellable_sql = text(f"""
        SELECT
            s.sku_code,
            s.description,
            s.long_description,
            s.cost_price,
            s.product_type,
            s.form_factor,
            s.attributes,
            s.use_stock,
            s.block_sales,
            s.gender,
            s.age_group,
            s.legacy_code,
            s.sale_ready,
            s.status,
            pr.price_incl_tax     AS retail_price,
            pr.price_excl_tax,
            pl.plu_code           AS nec_plu
        FROM skus s
        JOIN brands b ON b.id = s.brand_id
        LEFT JOIN prices pr
            ON pr.sku_id = s.id
            AND pr.valid_from <= CURRENT_DATE
            AND pr.valid_to   >= CURRENT_DATE
        LEFT JOIN plus pl ON pl.sku_id = s.id
        LEFT JOIN stores st ON st.id = s.store_id
        WHERE b.name = :brand_name
          AND s.status = 'active'
          AND s.block_sales = false
          {sellability_clause}
          {store_clause}
        ORDER BY s.sku_code
    """)

    excluded_sql = text("""
        SELECT
            s.sku_code,
            s.description,
            s.form_factor,
            s.sale_ready,
            s.status,
            s.block_sales,
            CASE WHEN pr.id IS NOT NULL THEN true ELSE false END AS has_price,
            CASE WHEN pl.id IS NOT NULL THEN true ELSE false END AS has_plu
        FROM skus s
        JOIN brands b ON b.id = s.brand_id
        LEFT JOIN prices pr
            ON pr.sku_id = s.id
            AND pr.valid_from <= CURRENT_DATE
            AND pr.valid_to   >= CURRENT_DATE
        LEFT JOIN plus pl ON pl.sku_id = s.id
        WHERE b.name = :brand_name
          AND NOT (
              s.sale_ready = true
              AND s.status = 'active'
              AND s.block_sales = false
              AND pr.id IS NOT NULL
              AND pl.id IS NOT NULL
              AND s.description IS NOT NULL
              AND s.description != ''
          )
        ORDER BY s.sku_code
    """)

    params: dict[str, Any] = {"brand_name": brand_name}
    if store_code:
        params["store_code"] = store_code

    async with engine.connect() as conn:
        result = await conn.execute(sellable_sql, params)
        sellable = [dict(row._mapping) for row in result]

        excl_result = await conn.execute(excluded_sql, {"brand_name": brand_name})
        excluded = [dict(row._mapping) for row in excl_result]

    await engine.dispose()
    return sellable, excluded


# ── Gap / exclusion reports ───────────────────────────────────────────────────

def print_data_gaps(products: list[dict[str, Any]]) -> None:
    """Report products in the export set that are still missing optional data."""
    no_cost = [p for p in products if not p.get("cost_price")]
    no_plu = [p for p in products if not (p.get("nec_plu") or p.get("plu_code"))]
    no_price = [p for p in products if not (p.get("retail_price") or p.get("price_incl_tax"))]

    if no_price:
        print(f"\n  WARNING  MISSING RETAIL PRICE: {len(no_price)} products (no PRICE row)")
    if no_plu:
        print(f"  WARNING  MISSING PLU/BARCODE:  {len(no_plu)} products (no PLU row)")
    if no_cost:
        print(f"  WARNING  MISSING COST PRICE:   {len(no_cost)} products")


def print_excluded_report(excluded: list[dict[str, Any]]) -> None:
    if not excluded:
        return
    not_sale_ready = [e for e in excluded if not e.get("sale_ready")]
    no_price = [e for e in excluded if not e.get("has_price")]
    no_plu = [e for e in excluded if not e.get("has_plu")]

    print(f"\n  EXCLUDED from NEC POS export: {len(excluded)} SKUs")
    if not_sale_ready:
        print(f"    sale_ready=false : {len(not_sale_ready)}")
    if no_price:
        print(f"    no active price  : {len(no_price)}")
    if no_plu:
        print(f"    no PLU/barcode   : {len(no_plu)}")

    if not_sale_ready:
        print(f"\n    Not sale-ready (first 10):")
        for e in not_sale_ready[:10]:
            flags = []
            if not e.get("has_price"):
                flags.append("no-price")
            if not e.get("has_plu"):
                flags.append("no-plu")
            flag_str = f"  [{', '.join(flags)}]" if flags else ""
            print(f"      {e['sku_code']:16s}  {(e.get('description') or '')[:45]}{flag_str}")


# ── JSON loader (legacy) ──────────────────────────────────────────────────────

def load_from_json(
    input_path: Path,
    location_filter: str = "",
    sale_ready_only: bool = True,
) -> list[dict[str, Any]]:
    data = json.loads(input_path.read_text())
    all_products = data.get("products", [])
    print(f"Loaded {len(all_products)} products from {input_path.name}")

    products = all_products
    if location_filter:
        products = [p for p in products
                    if p.get("stocking_location", "").lower() == location_filter.lower()]
        print(f"  Filtered by location='{location_filter}': {len(products)} products")
    if sale_ready_only:
        products = [p for p in products if p.get("sale_ready")]
        print(f"  Filtered by sale_ready=True: {len(products)} products")

    return products


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Export Jewel NEC POS Master Data (Excel)")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output Excel file path")
    parser.add_argument(
        "--from-json", action="store_true",
        help="Read from JSON master list instead of DB (legacy mode)"
    )
    parser.add_argument("--input", default=DEFAULT_INPUT, help="[JSON mode] Master product list JSON")
    parser.add_argument("--location", default="", help="[JSON mode] Filter by stocking_location")
    parser.add_argument("--brand", default=BRAND_NAME, help="[DB mode] Brand name to export")
    parser.add_argument("--store", default=None, help="[DB mode] Store code filter (e.g. JEWEL-01)")
    parser.add_argument(
        "--include-drafts", action="store_true",
        help="[DB mode] Include all active non-blocked SKUs, bypassing sale_ready gate"
    )
    parser.add_argument("--db-url", default=DEFAULT_DATABASE_URL, help="[DB mode] Database URL")
    args = parser.parse_args()

    def _repo_path(p: str) -> Path:
        path = Path(p)
        return path.resolve() if path.is_absolute() else (REPO_ROOT / p).resolve()

    output_path = _repo_path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Load products ──────────────────────────────────────────────────────
    excluded: list[dict[str, Any]] = []

    if args.from_json:
        input_path = _repo_path(args.input)
        products = load_from_json(input_path, args.location, sale_ready_only=not args.include_drafts)
        print(f"  [JSON mode] {len(products)} products for export")
    else:
        print(f"Connecting to database...")
        print(f"  Brand  : {args.brand}")
        if args.store:
            print(f"  Store  : {args.store}")
        mode_label = "DRAFT (bypassing sale_ready)" if args.include_drafts else "PRODUCTION (sale_ready=true required)"
        print(f"  Mode   : {mode_label}")

        try:
            products, excluded = asyncio.run(fetch_sellable_skus(
                database_url=args.db_url,
                brand_name=args.brand,
                store_code=args.store,
                include_drafts=args.include_drafts,
            ))
        except Exception as e:
            print(f"\nERROR connecting to database: {e}")
            print(f"\nIs cloud-sql-proxy running on port 5434?")
            print(f"  cloud-sql-proxy <project>:asia-southeast1:retailsg-db --port 5434")
            print(f"\nOr use JSON mode (no DB required):")
            print(f"  python export_nec_jewel.py --from-json")
            sys.exit(1)

        print(f"\n  Sellable SKUs found: {len(products)}")

    if not products:
        print("\n  No sellable products to export.")
        if excluded:
            print_excluded_report(excluded)
        return

    # ── Build workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_catg_sheet(wb)
    sku_count = build_sku_sheet(wb, products)
    plu_count = build_plu_sheet(wb, products)
    price_count = build_price_sheet(wb, products)
    promo_count = build_promo_sheet(wb, products)
    build_invdetails_sheet(wb)

    wb.save(str(output_path))

    # ── Summary ─────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  JEWEL NEC POS MASTER DATA EXPORT")
    print(f"{'='*60}")
    print(f"  Output    : {output_path}")
    print(f"  Generated : {time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Source    : {'JSON (legacy)' if args.from_json else 'Database (live)'}")
    print(f"")
    print(f"  CATG sheet       : {len(TENANT_CATG_TREE)} category nodes")
    print(f"  SKU v2 sheet     : {sku_count} products")
    print(f"  PLU sheet        : {plu_count} barcode mappings")
    print(f"  PRICE sheet      : {price_count} priced items")
    print(f"  PROMO sheet      : {promo_count} rows ({len(PROMO_TIERS)} tiers x {sku_count} SKUs)")
    for disc_id, label, _pct in PROMO_TIERS:
        print(f"                       {disc_id:22s} {label}")
    print(f"  INVDETAILS sheet : (empty template)")

    if price_count < sku_count:
        print(f"\n  WARNING  {sku_count - price_count} products have no PRICE row")

    print_data_gaps(products)
    if excluded:
        print_excluded_report(excluded)

    print(f"\n{'='*60}")
    print(f"  Ready to send to Jewel NEC POS team for import.")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
