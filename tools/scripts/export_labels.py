#!/usr/bin/env python3
"""
Export a labels spreadsheet for Brother P-touch Editor (Mac).

Generates an Excel workbook where each row = one label. Open the file in
Brother P-touch Editor, use "Link" (Database merge) to bind columns to
fields in your label template, and print as many labels as needed.

Default output: ``data/exports/jewel_labels.xlsx``

Usage:
  # Default — pull sellable SKUs from the live DB (cloud-sql-proxy on :5434)
  python export_labels.py

  # JSON fallback (no DB required)
  python export_labels.py --from-json

  # Pillar filter
  python export_labels.py --pillar homeware
  python export_labels.py --pillar jewellery
  python export_labels.py --pillar minerals

  # One row per physical unit at the target store (qty-aware)
  python export_labels.py --per-unit --inv-store JEWEL-01

  # Include all SKUs that have a PLU/barcode, not just sale-ready ones
  python export_labels.py --all-with-plu

  # Only products sourced from a specific supplier (by supplier_item_code)
  python export_labels.py --from-json --supplier hengweicraft --all-with-plu
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = "data/master_product_list.json"
DEFAULT_OUTPUT = "data/exports/jewel_labels.xlsx"
SUPPLIERS_DIR = "docs/suppliers"
DEFAULT_DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql+asyncpg://retailsg:RetailSG2026Secure@127.0.0.1:5434/retailsg",
)

BRAND_NAME = "VICTORIA ENSO"

# ── Catalogue pillars ────────────────────────────────────────────────────────
# Mirrors CAG_CATG_MAP in export_nec_jewel.py and data-quality-types.ts
HOMEWARE_TYPES = {
    "Figurine", "Sculpture", "Bookend", "Bowl", "Vase", "Box", "Tray",
    "Decorative Object", "Wall Art", "Gift Set", "Repair Service",
}
JEWELLERY_TYPES = {
    "Bracelet", "Necklace", "Ring", "Earring", "Charm", "Pendant",
    "Bead Strand", "Accessory",
}
MINERALS_TYPES = {
    "Loose Gemstone", "Raw Specimen", "Crystal Cluster", "Crystal Point",
    "Tumbled Stone", "Gemstone Bead", "Healing Crystal",
}

PILLAR_MAP = {
    "homeware":  HOMEWARE_TYPES,
    "jewellery": JEWELLERY_TYPES,
    "minerals":  MINERALS_TYPES,
}

# CAG category label per product_type (mirrors export_nec_jewel.py)
CAG_CATEGORY_LABEL: dict[str, str] = {
    "Bracelet":         "BANGLE & BRACELETS",
    "Necklace":         "NECKLACE",
    "Ring":             "RINGS",
    "Earring":          "EARRINGS",
    "Charm":            "CHARMS",
    "Pendant":          "CHARMS",
    "Bead Strand":      "COSTUME JEWELLERY",
    "Accessory":        "JEWELLERY ACCESSORY",
    "Loose Gemstone":   "PRECIOUS STONE/GOLD",
    "Raw Specimen":     "PRECIOUS STONE/GOLD",
    "Crystal Cluster":  "PRECIOUS STONE/GOLD",
    "Crystal Point":    "PRECIOUS STONE/GOLD",
    "Tumbled Stone":    "PRECIOUS STONE/GOLD",
    "Gemstone Bead":    "PRECIOUS STONE/GOLD",
    "Healing Crystal":  "PRECIOUS STONE/GOLD",
    "Figurine":          "DECORATIVE ITEM",
    "Sculpture":         "DECORATIVE ITEM",
    "Bookend":           "DECORATIVE ITEM",
    "Bowl":              "DECORATIVE ITEM",
    "Vase":              "DECORATIVE ITEM",
    "Box":               "DECORATIVE ITEM",
    "Tray":              "DECORATIVE ITEM",
    "Decorative Object": "DECORATIVE ITEM",
    "Wall Art":          "DECORATIVE ITEM",
    "Gift Set":          "GENERAL SOUVENIRS",
    "Repair Service":    "REPAIR SERVICE",
}


# ── Label-row builder ────────────────────────────────────────────────────────

GENERIC_PRODUCT_TYPES = {"finished", "material", "manufactured"}

ITEM_LABEL_COLUMNS = [
    "SKU_CODE",        # our internal code (e.g. VEBKECRYS0000067)
    "BARCODE",         # NEC PLU — drives the printed barcode symbol
    "BRAND",           # "VICTORIA ENSO"
    "DESCRIPTION",     # short product name (<=40 chars)
    "MATERIAL",        # primary material
    "PRODUCT_TYPE",    # e.g. "Sculpture"
    "CAG_CATEGORY",    # e.g. "DECORATIVE ITEM"
    "PRICE",           # formatted "$120"
    "PRICE_NUMERIC",   # raw number, for conditional P-touch logic
    "QTY",             # labels to print on this row (usually 1)
    "INTERNAL_CODE",   # supplier code, for staff reference
]

BOX_LABEL_COLUMNS = [
    "SKU_CODE",
    "BARCODE",
    "BRAND",
    "DESCRIPTION",
    "MATERIAL",
    "PRODUCT_TYPE",
    "QTY_IN_BOX",      # qty_on_hand — how many units are in this box
    "UNIT_PRICE",      # formatted "$120" (may be blank)
    "TOTAL_VALUE",     # formatted "$600" (qty × unit_price, blank if no price)
    "COST_PRICE",      # raw SGD cost for internal visibility
    "LOCATION",        # stocking_location e.g. "breeze_by_the_east"
    "SUPPLIER_CODE",   # supplier item code (internal_code)
]

# Backwards compat alias (was LABEL_COLUMNS before the box/item split)
LABEL_COLUMNS = ITEM_LABEL_COLUMNS


def _material(p: dict[str, Any]) -> str:
    attrs = p.get("attributes") or {}
    if isinstance(attrs, str):
        try:
            attrs = json.loads(attrs)
        except Exception:
            attrs = {}
    return (
        attrs.get("materials")
        or attrs.get("material")
        or p.get("material")
        or ""
    )[:40]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.upper() in {"", "NULL", "NONE"}:
        return ""
    return text


def _catalogue_type(p: dict[str, Any]) -> str:
    """Return the human-facing product category used for labels and pillar filters.

    In Postgres, ``product_type`` stores a generic enum and ``form_factor`` stores
    the actual category like ``Bookend`` or ``Sculpture``. In JSON mode,
    ``product_type`` already contains that category. Prefer ``form_factor`` when it
    exists so both sources resolve to the same label type.
    """
    form_factor = str(p.get("form_factor") or "").strip()
    product_type = str(p.get("product_type") or "").strip()
    if form_factor:
        return form_factor
    if product_type in GENERIC_PRODUCT_TYPES:
        return ""
    return product_type


def _item_label_row(p: dict[str, Any], qty_override: int | None = None) -> list[Any]:
    """Build one ITEM-label row (small tag per unit)."""
    sku_code = (p.get("sku_code") or "")[:32]
    barcode = str(p.get("nec_plu") or p.get("plu_code") or "")
    desc = (p.get("description") or "")[:40]
    material = _material(p)
    ptype = _catalogue_type(p)
    cag = CAG_CATEGORY_LABEL.get(ptype, "")
    raw_price = p.get("retail_price") or p.get("price_incl_tax")
    price_num = float(raw_price) if raw_price is not None else None
    price_str = f"${price_num:.0f}" if price_num is not None else ""
    qty = qty_override if qty_override is not None else 1
    internal = _clean_text(p.get("internal_code") or p.get("legacy_code"))

    return [
        sku_code, barcode, BRAND_NAME, desc, material, ptype, cag,
        price_str, price_num, qty, internal,
    ]

# Backwards-compat alias (older name used by tests/other scripts).
_label_row = _item_label_row


def _box_label_row(p: dict[str, Any]) -> list[Any]:
    """Build one BOX-label row (one per SKU — labels a storage/shipping box)."""
    sku_code = (p.get("sku_code") or "")[:32]
    barcode = str(p.get("nec_plu") or p.get("plu_code") or "")
    desc = (p.get("description") or "")[:40]
    material = _material(p)
    ptype = _catalogue_type(p)
    qty = int(p.get("qty_on_hand") or 0)
    raw_price = p.get("retail_price") or p.get("price_incl_tax")
    price_num = float(raw_price) if raw_price is not None else None
    price_str = f"${price_num:.0f}" if price_num is not None else ""
    total_val = (price_num * qty) if (price_num is not None and qty) else None
    total_val_str = f"${total_val:.0f}" if total_val is not None else ""
    cost = p.get("cost_price")
    cost_num = float(cost) if cost is not None else None
    location = _clean_text(p.get("stocking_location") or p.get("primary_stocking_location"))
    supplier_code = _clean_text(p.get("internal_code") or p.get("legacy_code"))

    return [
        sku_code, barcode, BRAND_NAME, desc, material, ptype,
        qty, price_str, total_val_str, cost_num, location, supplier_code,
    ]


# ── Workbook styling ─────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")


def _style_header(ws: Any) -> None:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _auto_width(ws: Any, max_width: int = 40) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = len(str(ws.cell(row=1, column=col).value or ""))
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def build_labels_workbook(
    wb: openpyxl.Workbook,
    products: list[dict[str, Any]],
    per_unit: bool = False,
    include_box: bool = True,
) -> tuple[int, int]:
    """Build ItemLabels (+ optionally BoxLabels) sheets. Returns (item_rows, box_rows).

    ItemLabels: one row per physical unit if ``per_unit`` else one per SKU.
    BoxLabels : one row per SKU (the unit of 'a storage box'); skipped if
    ``include_box=False``.
    """
    # Drop the default blank sheet; we manage sheets explicitly.
    if wb.active and wb.active.title == "Sheet":
        wb.remove(wb.active)

    item_ws = wb.create_sheet("ItemLabels")
    item_ws.append(ITEM_LABEL_COLUMNS)
    item_rows = 0
    for p in products:
        if not p.get("sku_code") or not (p.get("nec_plu") or p.get("plu_code")):
            continue
        if per_unit:
            qty = int(p.get("qty_on_hand") or 0)
            if qty <= 0:
                continue
            for _ in range(qty):
                item_ws.append(_item_label_row(p, qty_override=1))
                item_rows += 1
        else:
            item_ws.append(_item_label_row(p))
            item_rows += 1
    _style_header(item_ws)
    _auto_width(item_ws)

    box_rows = 0
    if include_box:
        box_ws = wb.create_sheet("BoxLabels")
        box_ws.append(BOX_LABEL_COLUMNS)
        for p in products:
            if not p.get("sku_code") or not (p.get("nec_plu") or p.get("plu_code")):
                continue
            # A "box" only makes sense if there's at least one unit of stock.
            # Still emit a row if qty is 0 so user can tag an empty display box.
            box_ws.append(_box_label_row(p))
            box_rows += 1
        _style_header(box_ws)
        _auto_width(box_ws)
    return item_rows, box_rows


# Backwards-compat shim
def build_labels_sheet(wb: openpyxl.Workbook, products, per_unit: bool = False) -> int:
    item_rows, _ = build_labels_workbook(wb, products, per_unit=per_unit)
    return item_rows


# ── DB fetch ─────────────────────────────────────────────────────────────────

async def fetch_labelable_skus(
    database_url: str,
    brand_name: str = BRAND_NAME,
    inv_store_code: str = "JEWEL-01",
    all_with_plu: bool = False,
) -> list[dict[str, Any]]:
    """Query DB for SKUs that should get labels.

    By default: sellable SKUs only (sale_ready + active price + PLU + description).
    With ``all_with_plu=True``: any active non-blocked SKU that has a PLU/barcode.
    """
    try:
        from sqlalchemy.ext.asyncio import create_async_engine
        from sqlalchemy import text
    except ImportError:
        print("ERROR: sqlalchemy[asyncio] not installed.")
        sys.exit(1)

    engine = create_async_engine(database_url, echo=False)

    sellability_clause = "" if all_with_plu else """
          AND s.sale_ready = true
          AND pr.id IS NOT NULL
          AND s.description IS NOT NULL
          AND s.description != ''"""

    sql = text(f"""
        SELECT
            s.sku_code,
            s.description,
            s.product_type,
            s.form_factor,
            s.attributes,
            s.legacy_code               AS internal_code,
            s.cost_price,
            s.primary_stocking_location AS stocking_location,
            pr.price_incl_tax            AS retail_price,
            pl.plu_code                  AS nec_plu,
            COALESCE(inv.qty_on_hand, 0) AS qty_on_hand
        FROM skus s
        JOIN brands b ON b.id = s.brand_id
        LEFT JOIN prices pr
            ON pr.sku_id = s.id
            AND pr.valid_from <= CURRENT_DATE
            AND pr.valid_to   >= CURRENT_DATE
        LEFT JOIN plus pl ON pl.sku_id = s.id
        LEFT JOIN stores inv_st ON inv_st.store_code = :inv_store_code
        LEFT JOIN inventories inv
            ON inv.sku_id = s.id
            AND inv.store_id = inv_st.id
        WHERE b.name = :brand_name
          AND s.status = 'active'
          AND s.block_sales = false
          AND pl.id IS NOT NULL
          {sellability_clause}
        ORDER BY s.product_type, s.sku_code
    """)

    params = {"brand_name": brand_name, "inv_store_code": inv_store_code}
    async with engine.connect() as conn:
        result = await conn.execute(sql, params)
        rows = [dict(row._mapping) for row in result]
    await engine.dispose()
    return rows


# ── Supplier code loader ─────────────────────────────────────────────────────

def load_supplier_codes(supplier_name: str) -> tuple[set[str], Path]:
    """Collect every ``supplier_item_code`` from ``docs/suppliers/<name>/orders/*.json``.

    Returns (codes, supplier_dir). Raises ``FileNotFoundError`` if the supplier
    directory doesn't exist.
    """
    supplier_dir = REPO_ROOT / SUPPLIERS_DIR / supplier_name
    if not supplier_dir.is_dir():
        raise FileNotFoundError(f"Supplier directory not found: {supplier_dir}")
    orders_dir = supplier_dir / "orders"
    codes: set[str] = set()
    if orders_dir.is_dir():
        for order_file in sorted(orders_dir.glob("*.json")):
            try:
                data = json.loads(order_file.read_text())
            except Exception as e:
                print(f"  WARNING  could not parse {order_file.name}: {e}")
                continue
            for li in data.get("line_items", []):
                code = li.get("supplier_item_code")
                if code:
                    codes.add(code)
    return codes, supplier_dir


def filter_by_supplier_codes(
    products: list[dict[str, Any]],
    codes: set[str],
    supplier_name: str,
) -> list[dict[str, Any]]:
    """Keep only products whose ``internal_code`` matches a supplier code.

    Prints a report of matched/unmatched codes so the user can see gaps.
    """
    matched = [p for p in products if p.get("internal_code") in codes]
    matched_codes = {p.get("internal_code") for p in matched}
    missing = sorted(codes - matched_codes)
    print(f"\n  Supplier '{supplier_name}' codes: {len(codes)}")
    print(f"  Matched in master list : {len(matched)} / {len(codes)}")
    if missing:
        print(f"  NOT YET IN MASTER ({len(missing)}): {', '.join(missing[:12])}" + (" ..." if len(missing) > 12 else ""))
        print(f"  (these supplier codes need to be ingested before they can be labelled)")
    return matched


# ── JSON loader (legacy) ─────────────────────────────────────────────────────

def load_from_json(input_path: Path, all_with_plu: bool) -> list[dict[str, Any]]:
    data = json.loads(input_path.read_text())
    all_products = data.get("products", [])
    print(f"Loaded {len(all_products)} products from {input_path.name}")
    products = [p for p in all_products if p.get("nec_plu")]
    print(f"  With PLU/barcode: {len(products)}")
    if not all_with_plu:
        products = [p for p in products if p.get("sale_ready")]
        print(f"  Filtered to sale_ready: {len(products)}")
    return products


# ── CLI ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Export labels spreadsheet for Brother P-touch Editor")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output Excel file path")
    parser.add_argument(
        "--from-json", action="store_true",
        help="Read from JSON master list instead of DB (legacy mode)"
    )
    parser.add_argument("--input", default=DEFAULT_INPUT, help="[JSON mode] Master product list JSON")
    parser.add_argument("--brand", default=BRAND_NAME, help="[DB mode] Brand name")
    parser.add_argument(
        "--inv-store", default="JEWEL-01",
        help="[DB mode] Store whose qty_on_hand is used (default: JEWEL-01)"
    )
    parser.add_argument(
        "--pillar", choices=sorted(PILLAR_MAP.keys()),
        help="Only include a single catalogue pillar (homeware/jewellery/minerals)"
    )
    parser.add_argument(
        "--all-with-plu", action="store_true",
        help="Include every SKU with a PLU/barcode, not just sale-ready ones"
    )
    parser.add_argument(
        "--supplier", default=None,
        help="Filter to products from this supplier (reads docs/suppliers/<name>/orders/*.json and matches internal_code)"
    )
    parser.add_argument(
        "--per-unit", action="store_true",
        help="Emit one row per physical unit at --inv-store (uses qty_on_hand)"
    )
    parser.add_argument(
        "--no-box", action="store_true",
        help="Skip the BoxLabels sheet (item-only export)"
    )
    parser.add_argument(
        "--only-plus", default=None,
        help="Comma-separated list of PLU/barcode values to keep (matches nec_plu/plu_code)"
    )
    parser.add_argument("--db-url", default=DEFAULT_DATABASE_URL, help="[DB mode] Database URL")
    args = parser.parse_args()

    only_plus = (
        {p.strip() for p in args.only_plus.split(",") if p.strip()}
        if args.only_plus
        else None
    )

    def _repo_path(p: str) -> Path:
        path = Path(p)
        return path.resolve() if path.is_absolute() else (REPO_ROOT / p).resolve()

    output_path = _repo_path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Load products ──────────────────────────────────────────────────────
    # When --only-plus is given, bypass the sale_ready filter so the explicit
    # PLU set is always honoured (e.g. test labels for not-yet-priced SKUs).
    effective_all_with_plu = args.all_with_plu or bool(only_plus)
    if args.from_json:
        products = load_from_json(_repo_path(args.input), effective_all_with_plu)
    else:
        print(f"Connecting to database...")
        print(f"  Brand     : {args.brand}")
        print(f"  InvStore  : {args.inv_store}")
        print(f"  Mode      : {'ALL with PLU' if args.all_with_plu else 'Sale-ready only'}")
        try:
            products = asyncio.run(fetch_labelable_skus(
                database_url=args.db_url,
                brand_name=args.brand,
                inv_store_code=args.inv_store,
                all_with_plu=effective_all_with_plu,
            ))
        except Exception as e:
            print(f"\nERROR connecting to database: {e}")
            print(f"\nIs cloud-sql-proxy running on port 5434?")
            print(f"Or use JSON mode:  python export_labels.py --from-json")
            sys.exit(1)
        print(f"  SKUs with PLU: {len(products)}")

    # Explicit PLU allowlist (e.g. small test batch for P-touch labels)
    if only_plus:
        before = len(products)
        products = [
            p for p in products
            if str(p.get("nec_plu") or p.get("plu_code") or "") in only_plus
        ]
        matched = {str(p.get("nec_plu") or p.get("plu_code")) for p in products}
        missing = sorted(only_plus - matched)
        print(f"  --only-plus filter: {before} -> {len(products)} products")
        if missing:
            print(f"  PLUs not found in source: {', '.join(missing)}")

    # Supplier filter (by internal_code ↔ supplier_item_code)
    if args.supplier:
        try:
            codes, supplier_dir = load_supplier_codes(args.supplier)
        except FileNotFoundError as e:
            print(f"\nERROR: {e}")
            sys.exit(1)
        print(f"  Supplier dir : {supplier_dir.relative_to(REPO_ROOT)}")
        products = filter_by_supplier_codes(products, codes, args.supplier)

    # Pillar filter
    if args.pillar:
        pillar_types = PILLAR_MAP[args.pillar]
        before = len(products)
        products = [
            p for p in products
            if _catalogue_type(p) in pillar_types
        ]
        print(f"  Pillar '{args.pillar}': {before} -> {len(products)} products")

    if not products:
        print("\n  No labelable products.")
        return

    # ── Build workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    item_rows, box_rows = build_labels_workbook(
        wb, products, per_unit=args.per_unit, include_box=not args.no_box,
    )
    wb.save(str(output_path))

    # ── Summary ─────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  LABELS EXPORT")
    print(f"{'='*60}")
    print(f"  Output        : {output_path}")
    print(f"  SKUs          : {len(products)}")
    print(f"  ItemLabels    : {item_rows} rows  {'(one per physical unit)' if args.per_unit else '(one per SKU)'}")
    if args.no_box:
        print(f"  BoxLabels     : (skipped \u2014 --no-box)")
    else:
        print(f"  BoxLabels     : {box_rows} rows  (one per SKU \u2014 storage/shipping box tag)")
    print(f"\n  Open in Brother P-touch Editor -> File -> Database -> Connect")
    print(f"  Pick the 'ItemLabels' sheet for small product tags, 'BoxLabels' for larger box tags.")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
